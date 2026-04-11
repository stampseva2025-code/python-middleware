import threading
import time
from flask import Flask, request, jsonify
import cv2
import numpy as np
import base64
from deep_translator import GoogleTranslator
import requests
import re
import os

os.environ['KMP_DUPLICATE_OK'] = 'True'
os.environ['OMP_NUM_THREADS'] = '1'
import json
import google.generativeai as genai
from flask_cors import CORS
import imagehash
from PIL import Image
import io
from PIL import Image as PILImage
from sentence_transformers import SentenceTransformer, util
from ezodf import newdoc, opendoc, Sheet
from io import BytesIO
from PIL import Image, ImageOps
import subprocess
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from google.oauth2 import service_account
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.http import MediaIoBaseDownload


from sentence_transformers import SentenceTransformer, util
import torch
import faiss
import bcrypt
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
from flask_bcrypt import Bcrypt
import decimal
import datetime
import socket
from datetime import  timezone
from datetime import datetime
import datetime as dt


# . 1KwQhFMqBLkitbvOcuHycMhVSXS80JCpy MASTER_FOLDER_ID = '1KwQhFMqBLkitbvOcuHycMhVSXS80JCpy'


# 1. Setup Gemini
GOOGLE_API_KEY = ""
genai.configure(api_key=GOOGLE_API_KEY)
# model = genai.GenerativeModel('gemini-2.5-flash')
# This points to the stable version with the 1,500 daily limit
# Change your model initialization to this:
# model = genai.GenerativeModel('gemini-2.0-flash-lite')
# model = genai.GenerativeModel('gemini-flash-latest')
model = genai.GenerativeModel('gemma-3-27b-it')
# import google.generativeai as genai
# genai.configure(api_key=GOOGLE_API_KEY)
# for m in genai.list_models():
#     if 'generateContent' in m.supported_generation_methods:
#         print(m.name)

SERPAPI_API_KEY = ""

imgmodel = SentenceTransformer('clip-ViT-B-32')

app = Flask(__name__)
CORS(app)

# --- GOOGLE DRIVE CONFIG ---
SERVICE_ACCOUNT_FILE = 'service_account.json'
MASTER_FOLDER_ID = ''

CLIENT_SECRETS_FILE = 'client_secrets.json'
TOKEN_FILE = 'token.json'
SCOPES = ['https://www.googleapis.com/auth/drive.file']
LOCAL_BACKUP_DIR = '/Users/navinkumar/Public/projects/gdb' # <-- UPDATE THIS

PLAN_FILE = 'drive_sync_plan.json'
STATUS_FILE = 'drive_sync_status.json'

# --- INITIALIZE THEME MAPPING ---
# This loads a small 80MB model to understand word meanings
device = "mps" if torch.backends.mps.is_available() else "cpu"
theme_model = SentenceTransformer('all-MiniLM-L6-v2', device=device)

# YOUR EXACT 9 CATEGORIES
FIXED_CATEGORIES = [
    "Animal", "People", "Icons", "Rivals",
    "Landscape", "Flags", "Horse", "Festive", "Food"
]

# ==========================================
# 🗺️ 3. THE SYNC BLUEPRINT
# ==========================================
SYNC_BLUEPRINT = {
    "users": ["id", "name", "role", "password", "updated_at"],
    "stamp_tags": ["id", "stamp_id", "tag", "updated_at"],
    "stamp_images": ["id", "stamp_id", "image_url", "image_hash", "has_cancellation", "is_primary", "created_at",
                     "updated_at"],
    "sheets": ["id", "sheet_name", "project_prefix", "country", "issue_year", "description", "source", "uploadedBy",
               "uploadDate", "uploadTime", "status", "original_image_url", "drive_folder_id",
               "created_at", "v1_url", "v2_url", "v3_url", "v4_url", "v5_url", "updated_at"],
    "stamps": ["id", "sheet_id", "fileName", "folder", "imagePath", "Country", "THEME", "Year", "Color", "Denomination",
               "extra_copies", "initials", "estimated_value", "History", "Description", "historical_context",
               "curator_fun_fact", "design_symbolism", "narrative_script", "Remarks", "ai_raw", "fingerprint_phash",
               "fingerprint_dhash", "fingerprint_whash", "Operator", "created_at", "drive_url", "local_path",
               "drive_folder_id", "updated_at"],
    "duplicate_audit": ["stamp_id", "status", "dist_p", "dist_d", "dist_w", "original_id", "target_fileName",
                        "target_drive_url", "original_fileName", "original_drive_url", "user_resolution", "checked_at"]
}

#########################################################################
#########################################################################
#########################################################################
#########################################################################
#########################################################################
# 1. Read the environment variable (Defaults to 'cloud' if not found)
MODE = 'cloud'  # Default mode (cloud) disaster recovery mode dr
#########################################################################
#########################################################################
#########################################################################
#########################################################################
#########################################################################

# 2. Set the variables dynamically
if MODE == 'dr':
    CLOUD_API_URL = "https://laure-formfitting-cecil.ngrok-free.dev"
    SECRET_API_KEY = ""
    print("⚠️ WARNING: Running in DISASTER RECOVERY mode")
else:
    CLOUD_API_URL = "https://stampseva2025.pythonanywhere.com"
    SECRET_API_KEY = ""
    print("☁️ Running in NORMAL mode")

# Continue with your code...

# --- CONFIGURATION ---
CLOUD_AUDIT_URL = f"{CLOUD_API_URL}/audit-next-stamp"
HEADERS = {'X-API-KEY': SECRET_API_KEY, 'Content-Type': 'application/json', 'ngrok-skip-browser-warning': 'true'}

# --- GLOBAL STATE ---
audit_running = False
audit_thread = None
run_count = 0
start_time = None

# Pre-calculate category vectors for instant comparison
category_embeddings = theme_model.encode(FIXED_CATEGORIES, convert_to_tensor=True)

# Ensure you have your global faiss_index variable declared at the top of your file
faiss_index = None

# Replace 'global faiss_index' at the top of your file with a dictionary:
faiss_cache = {}


def get_vector_from_base64(base64_string):
    """
    Decodes a Base64 string into an image and generates its CLIP vector.
    """
    try:
        # 1. Strip the header if present (e.g., "data:image/jpeg;base64,...")
        if ',' in base64_string:
            base64_string = base64_string.split(',')[1]

        # 2. Decode the Base64 string into raw bytes
        image_data = base64.b64decode(base64_string)

        # 3. Convert raw bytes into a PIL Image object
        img = Image.open(io.BytesIO(image_data))

        # 4. Generate embedding (Vector)
        vector = model.encode(img)
        return vector

    except Exception as e:
        print(f"Error generating vector: {e}")
        return None


def periodic_auditor():
    global audit_running, run_count
    print("🚀 Periodic Auditor Thread Started...")

    while audit_running:
        try:
            response = requests.post(CLOUD_AUDIT_URL, headers=HEADERS, timeout=30)
            data = response.json()

            # --- DEBUG PRINTS START ---
            print(f"📡 Cloud API Response: {data.get('status')} | Full Data: {data}")
            # --- DEBUG PRINTS END ---

            if data.get('status') == 'success':
                run_count += 1
                # Additional notification for the Batch of 5 logic
                matches = data.get('matches_found', 0)
                print(f"[{time.strftime('%H:%M:%S')}] Audit #{run_count} completed for ID {data.get('audited_id')}")
                print(f"🔍 Result: {matches} potential duplicates identified and logged.")

            elif data.get('status') == 'complete':
                print("✅ All stamps audited. Stopping.")
                audit_running = False
                break

            # Wait 20 seconds before the next hit
            for _ in range(20):
                if not audit_running: break
                time.sleep(1)

        except Exception as e:
            print(f"❌ Connection Error: {e}")
            time.sleep(10)


def get_mapped_theme(ai_theme):
    if not ai_theme:
        return "Others"

    # Calculate "meaning" vector for the AI's theme
    ai_vector = theme_model.encode(ai_theme, convert_to_tensor=True)

    # Compare against your 9 categories
    scores = util.cos_sim(ai_vector, category_embeddings)[0]

    # Get highest similarity score
    best_idx = torch.argmax(scores).item()

    # Optional: If the match is extremely weak (meaningless text), return Others
    if scores[best_idx].item() < 0.30:
        return "Others"

    return FIXED_CATEGORIES[best_idx]


def compare_images_ai(base64_a, base64_b, model):
    # 1. Convert Base64 to Image Objects
    img1 = Image.open(BytesIO(base64.b64decode(base64_a.split(',')[1] if ',' in base64_a else base64_a)))
    img2 = Image.open(BytesIO(base64.b64decode(base64_b.split(',')[1] if ',' in base64_b else base64_b)))

    # 2. Encode both images to vectors
    vector1 = model.encode(img1)
    vector2 = model.encode(img2)

    # 3. Calculate Cosine Similarity
    score = util.cos_sim(vector1, vector2).item()

    return score * 100  # Returns percentage (e.g., 98.5%)


def compare_images_features(base64_a, base64_b):
    # 1. Decode Base64 to OpenCV format
    def b64_to_cv2(b64_str):
        b64_data = b64_str.split(',')[1] if ',' in b64_str else b64_str
        nparr = np.frombuffer(base64.b64decode(b64_data), np.uint8)
        return cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)

    img1 = b64_to_cv2(base64_a)
    img2 = b64_to_cv2(base64_b)

    # 2. Initialize ORB detector
    orb = cv2.ORB_create()

    # 3. Find keypoints and descriptors
    kp1, des1 = orb.detectAndCompute(img1, None)
    kp2, des2 = orb.detectAndCompute(img2, None)

    # 4. Match descriptors (BFMatcher)
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

    if des1 is None or des2 is None: return 0  # Comparison failed

    matches = bf.match(des1, des2)

    # 5. Calculate Score based on successful matches
    # This is a rough heuristic: more matches = better similarity
    matches = sorted(matches, key=lambda x: x.distance)

    # Take top 50 matches (or fewer if not enough found)
    top_matches = matches[:50]

    if len(top_matches) == 0: return 0

    # Calculate average distance (lower is better)
    avg_dist = sum(m.distance for m in top_matches) / len(top_matches)

    # Convert distance to a roughly 0-100 score (Distance 0 = 100% match)
    # 60 is a typical "bad" distance threshold for ORB
    similarity = max(0, 100 - avg_dist)

    return similarity


def get_standardized_img(image_b64):
    # 1. Decode
    img_data = base64.b64decode(image_b64)
    img = Image.open(io.BytesIO(img_data))

    # 2. Fix Orientation
    # This is non-negotiable for mobile.
    img = ImageOps.exif_transpose(img)

    # 3. Convert to Grayscale
    img = img.convert("L")

    # 4. Global Histogram Equalization
    # This is BETTER than Contrast(2.0). It expands the contrast
    # based on the image's own light levels, making phone and
    # laptop images look identical in terms of brightness.
    img = ImageOps.equalize(img)

    # 5. Fixed Resizing with LANCZOS
    # We use 256x256. It's high enough to keep stamp details
    # but small enough to wash out minor sensor noise.
    # LANCZOS is the most consistent across different OS platforms.
    img = img.resize((256, 256), Image.Resampling.LANCZOS)

    return img


def get_drive_service():
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)


def get_or_create_folder(folder_name, parent_id, service):
    query = f"name = '{folder_name}' and '{parent_id}' in parents and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    results = service.files().list(q=query, fields="files(id)").execute()
    items = results.get('files', [])
    if items:
        return items[0]['id']
    else:
        folder_metadata = {'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder',
                           'parents': [parent_id]}
        folder = service.files().create(body=folder_metadata, fields='id').execute()
        return folder.get('id')


def silent_drive_delete(file_id):
    """Internal helper to wipe files from Drive with detailed troubleshooting"""
    try:
        # 1. Extract the ID
        if not file_id:
            print(f"[DRIVE SERVICE] ❌ ERROR: Could not extract File ID from ID: {file_id}")
            return False

        print(f"[DRIVE SERVICE] 🔍 Extracted File ID: {file_id}")

        # 2. Connect to service
        service = get_drive_service()

        # 3. Perform Deletion
        print(f"[DRIVE SERVICE] 🗑️ Attempting to delete file ID: {file_id} from Google Drive...")
        service.files().delete(fileId=file_id).execute()

        print(f"[DRIVE SERVICE] ✨ SUCCESS: File {file_id} has been permanently removed.")
        return True

    except Exception as e:
        # If the file is already gone (404), we should log it specifically
        error_msg = str(e)
        if "File not found" in error_msg or "404" in error_msg:
            print(f"[DRIVE SERVICE] ℹ️ Skip: File {file_id} already deleted or does not exist.")
            return True  # Return true because the goal (file gone) is met
        else:
            print(f"[DRIVE SERVICE] ❌ CRITICAL ERROR: {error_msg}")
            return False


def remove_black_background_one(image_data):
    # 1. Decode
    nparr = np.frombuffer(base64.b64decode(image_data.split(',')[1]), np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)

    if img.shape[2] == 4:
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

    # 2. Focus on "Darkness"
    # We use the Value channel from HSV because it's best at seeing "Black vs Color"
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    v_channel = hsv[:, :, 2]

    # 3. OTSU'S AUTO-THRESHOLD
    # This automatically finds the best "cut-off" point for the black background
    _, mask = cv2.threshold(v_channel, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # 4. Cleanup Noise (Distortion)
    # This removes tiny black specks and smooths the "distorted" edges
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=2)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)

    # 5. FIND THE BIGGEST OBJECT
    # We ignore any "distorted" noise and only keep the largest shape (the stamp)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return img

    max_cnt = max(contours, key=cv2.contourArea)

    # Create a clean, smooth mask of just the stamp
    clean_mask = np.zeros_like(mask)
    cv2.drawContours(clean_mask, [max_cnt], -1, 255, -1)

    # 6. Apply Transparency
    b, g, r = cv2.split(img)
    rgba = cv2.merge([b, g, r, clean_mask])

    # 7. Final Tight Crop
    x, y, w, h = cv2.boundingRect(max_cnt)

    # SHAVE: Since you see distortion, we shave 8 pixels inside
    # to make sure the "edge noise" is completely deleted.
    shave = 8
    final_crop = rgba[y + shave: y + h - shave, x + shave: x + w - shave]

    return final_crop


def remove_black_background_two(image_data):
    # 1. Decode original image
    nparr = np.frombuffer(base64.b64decode(image_data.split(',')[1]), np.uint8)
    original_img = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)

    # Ensure BGR (remove alpha if present)
    if original_img.shape[2] == 4:
        clean_img = cv2.cvtColor(original_img, cv2.COLOR_BGRA2BGR)
    else:
        clean_img = original_img.copy()

    # 2. Convert to grayscale
    gray = cv2.cvtColor(clean_img, cv2.COLOR_BGR2GRAY)

    # 3. Blur to unify perforations into a continuous wall
    blurred = cv2.GaussianBlur(gray, (7, 7), 0)

    # 4. Adaptive threshold (robust for colored stamps)
    thresh = cv2.adaptiveThreshold(
        blurred,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        15,  # slightly larger window for safety
        3
    )

    # 5. Morphology: seal perforation gaps
    kernel = np.ones((3, 3), np.uint8)
    sealed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=3)

    # 6. Find external contours
    contours, _ = cv2.findContours(sealed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return clean_img

    # Largest contour = stamp + margin
    max_cnt = max(contours, key=cv2.contourArea)

    # 7. Create a solid, impermeable mask
    mask = np.zeros(gray.shape, dtype=np.uint8)
    cv2.fillPoly(mask, [max_cnt], 255)

    # -------------------------------------------------
    # OPTIONAL VERY GENTLE ENHANCEMENT (SAFE)
    # Comment out this block if you want ZERO enhancement
    # -------------------------------------------------

    lab = cv2.cvtColor(clean_img, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)

    # Gentle CLAHE (safe for stamps)
    clahe = cv2.createCLAHE(clipLimit=1.8, tileGridSize=(8, 8))
    l = clahe.apply(l)

    enhanced = cv2.merge((l, a, b))
    enhanced = cv2.cvtColor(enhanced, cv2.COLOR_LAB2BGR)

    # -------------------------------------------------

    # 8. Apply mask → transparent background
    b, g, r = cv2.split(enhanced)
    rgba = cv2.merge((b, g, r, mask))

    # 9. Tight crop with 1px safety margin
    x, y, w, h = cv2.boundingRect(max_cnt)
    pad = 1
    h_img, w_img = mask.shape

    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(w_img, x + w + pad)
    y2 = min(h_img, y + h + pad)

    final_crop = rgba[y1:y2, x1:x2]

    return final_crop


def remove_black_background_three(image_data):
    # 1. Decode original image
    nparr = np.frombuffer(base64.b64decode(image_data.split(',')[1]), np.uint8)
    original_img = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)

    # Ensure BGR (remove alpha if present)
    if original_img.shape[2] == 4:
        clean_img = cv2.cvtColor(original_img, cv2.COLOR_BGRA2BGR)
    else:
        clean_img = original_img.copy()

    # -------------------------------------------------
    # 2. LAB COLOR SPACE (KEY IMPROVEMENT)
    # -------------------------------------------------
    lab = cv2.cvtColor(clean_img, cv2.COLOR_BGR2LAB)
    L, A, B = cv2.split(lab)

    # B channel separates paper vs background very well
    blurred = cv2.GaussianBlur(B, (7, 7), 0)

    # -------------------------------------------------
    # 3. OTSU THRESHOLD (more stable than adaptive here)
    # -------------------------------------------------
    _, thresh = cv2.threshold(
        blurred,
        0,
        255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )

    # Ensure stamp area is white
    if np.mean(thresh) > 127:
        thresh = cv2.bitwise_not(thresh)

    # -------------------------------------------------
    # 4. MORPHOLOGY — seal perforations safely
    # -------------------------------------------------
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    sealed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)
    sealed = cv2.morphologyEx(sealed, cv2.MORPH_OPEN, kernel, iterations=1)

    # -------------------------------------------------
    # 5. FIND STAMP CONTOUR
    # -------------------------------------------------
    contours, _ = cv2.findContours(sealed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return clean_img

    max_cnt = max(contours, key=cv2.contourArea)

    # -------------------------------------------------
    # 6. CREATE IMPERMEABLE MASK
    # -------------------------------------------------
    mask = np.zeros(B.shape, dtype=np.uint8)
    cv2.fillPoly(mask, [max_cnt], 255)

    # -------------------------------------------------
    # OPTIONAL VERY GENTLE ENHANCEMENT (UNCHANGED)
    # -------------------------------------------------
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=1.8, tileGridSize=(8, 8))
    l = clahe.apply(l)
    enhanced = cv2.merge((l, a, b))
    enhanced = cv2.cvtColor(enhanced, cv2.COLOR_LAB2BGR)
    # -------------------------------------------------

    # -------------------------------------------------
    # 7. APPLY MASK → TRANSPARENT BACKGROUND
    # -------------------------------------------------
    b, g, r = cv2.split(enhanced)
    rgba = cv2.merge((b, g, r, mask))

    # -------------------------------------------------
    # 8. TIGHT CROP (SAFE)
    # -------------------------------------------------
    x, y, w, h = cv2.boundingRect(max_cnt)
    pad = 1
    h_img, w_img = mask.shape

    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(w_img, x + w + pad)
    y2 = min(h_img, y + h + pad)

    final_crop = rgba[y1:y2, x1:x2]

    return final_crop


def remove_black_background_four(image_data):
    # 1. Decode Base64
    nparr = np.frombuffer(base64.b64decode(image_data.split(',')[1]), np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)

    if img.shape[2] == 4:
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

    original = img.copy()

    # 2. Blur to stabilize perforations
    blurred = cv2.GaussianBlur(img, (5, 5), 0)

    # 3. Initial shape mask (brightness-based)
    gray = cv2.cvtColor(blurred, cv2.COLOR_BGR2GRAY)
    _, shape_mask = cv2.threshold(gray, 35, 255, cv2.THRESH_BINARY)

    contours, _ = cv2.findContours(shape_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return original

    max_cnt = max(contours, key=cv2.contourArea)

    solid_mask = np.zeros_like(shape_mask)
    cv2.fillPoly(solid_mask, [max_cnt], 255)

    # 4. Seal perforations (shape only)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    solid_mask = cv2.morphologyEx(solid_mask, cv2.MORPH_CLOSE, kernel, iterations=2)

    # ----------------------------------------------------
    # 🔥 NEW: REMOVE NEAR-BLACK BACKGROUND BLEED
    # ----------------------------------------------------

    hsv = cv2.cvtColor(original, cv2.COLOR_BGR2HSV)

    # Define "background black" very strictly
    lower_black = np.array([0, 0, 0])
    upper_black = np.array([180, 60, 50])  # allows dark gray but NOT ink

    black_bg = cv2.inRange(hsv, lower_black, upper_black)

    # Remove black ONLY outside the stamp
    outside_mask = cv2.bitwise_not(solid_mask)
    black_outside = cv2.bitwise_and(black_bg, outside_mask)

    # Clean thin halos
    black_outside = cv2.morphologyEx(
        black_outside,
        cv2.MORPH_DILATE,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
        iterations=1
    )

    # Subtract background-black from solid mask
    final_mask = cv2.bitwise_and(
        solid_mask,
        cv2.bitwise_not(black_outside)
    )

    # ----------------------------------------------------
    # 5. Alpha compose
    # ----------------------------------------------------
    b, g, r = cv2.split(original)
    rgba = cv2.merge([b, g, r, final_mask])

    # ----------------------------------------------------
    # 6. Tight crop (safe)
    # ----------------------------------------------------
    coords = cv2.findNonZero(final_mask)
    if coords is not None:
        x, y, w, h = cv2.boundingRect(coords)
        pad = 1
        rgba = rgba[
            max(0, y - pad): y + h + pad,
            max(0, x - pad): x + w + pad
        ]

    return rgba


def remove_black_background_five(image_data):
    # 1. Decode the base64 image data
    nparr = np.frombuffer(base64.b64decode(image_data.split(',')[1]), np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("Failed to decode image")
    original = img.copy()

    # 2. Convert to LAB color space (better separation of lightness from color)
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    l_channel = lab[:, :, 0]  # L channel: lightness (0=black, 255=white)

    # 3. Create mask for very dark areas (black background + heavy cancellations)
    # L < 50 captures near-black regions (adjustable)
    dark_mask = cv2.inRange(l_channel, 0, 50)

    # 4. Invert to get potential stamp areas (higher lightness = paper + colors)
    potential_stamp = cv2.bitwise_not(dark_mask)

    # 5. Aggressive morphology to connect stamp parts and close gaps
    kernel_close = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (21, 21))
    closed = cv2.morphologyEx(potential_stamp, cv2.MORPH_CLOSE, kernel_close, iterations=5)

    # Clean small noise/holes
    kernel_open = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    cleaned = cv2.morphologyEx(closed, cv2.MORPH_OPEN, kernel_open, iterations=2)

    # 6. Find contours and select the largest one (the stamp)
    contours, _ = cv2.findContours(cleaned, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return original  # fallback

    max_cnt = max(contours, key=cv2.contourArea)

    # Safety: if largest contour is too small (<25% of image), fallback
    if cv2.contourArea(max_cnt) < 0.25 * img.shape[0] * img.shape[1]:
        return original

    # 7. Create precise stamp mask
    stamp_mask = np.zeros(img.shape[:2], dtype=np.uint8)
    cv2.fillPoly(stamp_mask, [max_cnt], 255)

    # 8. Soften edges for natural anti-aliased look
    stamp_mask = cv2.GaussianBlur(stamp_mask, (7, 7), 0)

    # 9. Apply as alpha channel
    b, g, r = cv2.split(original)
    rgba = cv2.merge([b, g, r, stamp_mask])

    # 10. Tight crop with generous padding to keep all perforations
    x, y, w, h = cv2.boundingRect(max_cnt)
    pad = 15  # increased padding
    cropped = rgba[
        max(0, y - pad): min(rgba.shape[0], y + h + pad),
        max(0, x - pad): min(rgba.shape[1], x + w + pad)
    ]

    return cropped


def straighten_stamp(img_cv):
    # 1. Handle Alpha/Transparency
    if img_cv.shape[2] == 4:
        _, mask = cv2.threshold(img_cv[:, :, 3], 1, 255, cv2.THRESH_BINARY)
    else:
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        _, mask = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # 2. Get Coordinates
    coords = cv2.findNonZero(mask)
    if coords is None or len(coords) < 5:
        return img_cv

    # 3. Get the Rotated Rect
    rect = cv2.minAreaRect(coords)
    (center), (w, h), angle = rect

    # 4. THE ASPECT RATIO PROTECTOR
    # OpenCV's angle logic changes depending on if width is > or < than height.
    # We want to ensure a 'tall' stamp stays 'tall'.

    actual_angle = angle
    if w < h:
        # If the rectangle is taller than it is wide
        actual_angle = angle
    else:
        # If the rectangle is wider than it is tall, OpenCV is misidentifying the 'top'
        actual_angle = angle + 90

    # 5. LIMIT ROTATION (The 'Anti-Flip' Guard)
    # We only want to fix tilts, not perform 90-degree rotations.
    # This forces the adjustment to be a small number (like 2 degrees, not 92).
    while actual_angle > 45:
        actual_angle -= 90
    while actual_angle < -45:
        actual_angle += 90

    # 6. Rotate
    (H, W) = img_cv.shape[:2]
    M = cv2.getRotationMatrix2D((W // 2, H // 2), actual_angle, 1.0)

    rotated = cv2.warpAffine(img_cv, M, (W, H),
                             flags=cv2.INTER_CUBIC,
                             borderMode=cv2.BORDER_CONSTANT,
                             borderValue=(0, 0, 0, 0))

    # 7. Final Clean Crop
    if rotated.shape[2] == 4:
        new_alpha = rotated[:, :, 3]
        new_coords = cv2.findNonZero(new_alpha)
        if new_coords is not None:
            x, y, wb, hb = cv2.boundingRect(new_coords)
            rotated = rotated[y:y + hb, x:x + wb]

    return rotated


def remove_cancellation_ink(image_data):
    # Decode image
    nparr = np.frombuffer(base64.b64decode(image_data.split(',')[1]), np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 1. Detect dark regions
    _, dark_mask = cv2.threshold(gray, 60, 255, cv2.THRESH_BINARY_INV)

    # 2. Remove thin lines (printed ink stays)
    kernel_open = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    opened = cv2.morphologyEx(dark_mask, cv2.MORPH_OPEN, kernel_open, iterations=1)

    # 3. Keep only large smooth regions (cancellation marks)
    kernel_close = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (9, 9))
    cancellation_mask = cv2.morphologyEx(opened, cv2.MORPH_CLOSE, kernel_close, iterations=2)

    # 4. Remove small components (protect text & line art)
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(cancellation_mask, connectivity=8)

    refined_mask = np.zeros_like(cancellation_mask)
    for i in range(1, num_labels):
        area = stats[i, cv2.CC_STAT_AREA]
        if area > 400:  # cancellation ink is usually large
            refined_mask[labels == i] = 255

    # 5. Inpaint ONLY those regions
    cleaned = cv2.inpaint(img, refined_mask, inpaintRadius=3, flags=cv2.INPAINT_TELEA)

    return cleaned


# Helper to extract File ID from your specific URL format
def extract_file_id(url):
    if not url:
        return None

    # 1. Look for the ID after 'picture/' followed by any number
    # This regex looks for a string of 25+ characters (standard Drive ID length)
    match = re.search(r'picture/\d+([a-zA-Z0-9_-]{25,})', url)
    if match:
        return match.group(1)

    # 2. Fallback: If your ID starts immediately after picture/ without a number
    match = re.search(r'picture/([a-zA-Z0-9_-]{25,})', url)
    if match:
        return match.group(1)

    # 3. Fallback: Standard Google Drive 'id=' format
    match = re.search(r'id=([a-zA-Z0-9_-]{25,})', url)
    if match:
        return match.group(1)

    return None


def new_extract_file_id(url):
    if not url or not isinstance(url, str):
        return None

    print(f"[EXTRACTOR] 🧪 Analyzing: {url}")

    # 1. Check if the URL is too short to be a real Drive link
    # If the URL ends in a single digit (like /picture/2), it's not a file
    if len(url.split('/')[-1]) < 10:
        print(f"[EXTRACTOR] ⚠️ Skipping: URL segment '{url.split('/')[-1]}' is too short to be a File ID.")
        return None

    # 2. Pattern for googleusercontent (Handles the '0' prefix and 33-char IDs)
    # This specifically looks for the long alphanumeric string after 'picture/'
    match = re.search(r'picture/(?:0|)?([a-zA-Z0-9_-]{25,})', url)
    if match:
        return match.group(1)

    # 3. Pattern for standard Google Drive links (id=...)
    match_id = re.search(r'id=([a-zA-Z0-9_-]{25,})', url)
    if match_id:
        return match_id.group(1)

    # 4. Pattern for shared links (d/FILE_ID/view)
    match_d = re.search(r'd/([a-zA-Z0-9_-]{25,})', url)
    if match_d:
        return match_d.group(1)

    return None


def upload_base64_to_freeimage(image_data):
    """
    Uploads base64 image data to freeimage.host anonymously (public guest key).
    Returns direct permanent URL.
    """
    try:
        base64_string = image_data.split(',')[1]
    except IndexError:
        print("Invalid data URL format")
        return None

    try:
        img_bytes = base64.b64decode(base64_string)
    except Exception as e:
        print(f"Base64 decode failed: {e}")
        return None

    upload_url = "https://freeimage.host/api/1/upload"

    # Public anonymous guest key (official from their docs)
    data = {
        "key": "6d207e02198a847aa98d0a2a901485a5",
        "action": "upload",
        "format": "json"
    }

    files = {
        "source": ("stamp.png", img_bytes, "image/png")
    }

    try:
        response = requests.post(upload_url, data=data, files=files, timeout=30)
        response.raise_for_status()

        result = response.json()

        if result.get("status_code") == 200:
            direct_url = result["image"]["url"]
            print(f"Upload successful! Direct URL: {direct_url}")
            return direct_url
        else:
            print(f"Upload error: {result}")
            return None

    except requests.exceptions.RequestException as e:
        print(f"Upload failed: {e}")
        return None


SERPAPI_ENDPOINT = "https://serpapi.com/search"
TIMEOUT = 30


def run_reverse_image_search(image_url: str, api_key: str):
    try:
        params = {
            "engine": "google_reverse_image",  # updated to Google reverse image search
            "image_url": image_url,
            "api_key": api_key
        }

        response = requests.get(
            SERPAPI_ENDPOINT,
            params=params,
            timeout=TIMEOUT
        )

        # HTTP error
        if response.status_code != 200:
            return {
                "success": False,
                "error": f"HTTP {response.status_code}: {response.text}"
            }

        data = response.json()

        # SerpApi error
        if "error" in data:
            return {
                "success": False,
                "error": data["error"]
            }

        return {
            "success": True,
            "data": data
        }

    except requests.exceptions.Timeout:
        return {
            "success": False,
            "error": "Request timed out"
        }

    except requests.exceptions.RequestException as e:
        return {
            "success": False,
            "error": str(e)
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"Unexpected error: {str(e)}"
        }


# --- Logic Function ---
def get_stamp_details(base64_image):
    # 1. Process Image for Fingerprinting
    if "," in base64_image:
        base64_image = base64_image.split(",")[1]

    image_bytes = base64.b64decode(base64_image)
    pil_img = Image.open(io.BytesIO(image_bytes))
    phash_val = str(imagehash.phash(pil_img))
    # --- ADDED: Calculate dHash ---
    dhash_val = str(imagehash.dhash(pil_img))
    # --- ADDED: Calculate wHash ---
    whash_val = str(imagehash.whash(pil_img))

    # 2. Maximum Museum Information Prompt
    # This includes your specific fields + the storytelling fields
    prompt = """
    Identify this postage stamp and return ONLY a JSON object. 
    Use these exact keys:

    - "Country": (The issuing nation)
    - "Year": (Year of issue)
    - "Denomination": (Face value as printed on the stamp)
    - "Color": (Dominant colors of the stamp)
    - "THEME": (Broad category like Birds, Flowers, Space, Royalty, etc.)
    - "estimated_Value": (Market value estimate in USD)
    - "Extra_Copies": (Set to "0" by default)
    - "Initials": (Set to "None" by default)
    - "Remarks": (Any notable defects, variations, or interesting physical notes)

    -- Storytelling Fields for Museum App --
    - "historical_context": (The world events during this stamp's era)
    - "design_symbolism": (Explain hidden meanings in the artwork)
    - "narrative_script": (A 30-second first-person script for an audio guide)
    - "curator_fun_fact": (A surprising trivia point for visitors)
    - "theme_tags": (4-5 SEO tags for your app's search engine)
    """

    image_part = {'mime_type': 'image/png', 'data': base64_image}
    response = model.generate_content([prompt, image_part])

    # 3. Parse and Clean the JSON
    clean_json = response.text.replace('```json', '').replace('```', '').strip()
    stamp_data = json.loads(clean_json)

    # 4. Attach the Visual Fingerprint for Duplicate Detection
    stamp_data["fingerprint_phash"] = phash_val
    # --- ADDED: Return dHash with the result ---
    stamp_data["fingerprint_dhash"] = dhash_val
    # --- ADDED: Return wHash with the result ---
    stamp_data["fingerprint_whash"] = whash_val

    print(jsonify({"success": True, "data": stamp_data}))

    return stamp_data

############################################################################
##############GOOGLE DRIVE FOLDER SCANNER & BACKUP CODE (STARTS)############
############################################################################

# def test_permissions():
#     try:
#         service = authenticate_drive()
#         # We try to look directly at the Master Folder, not what's inside it
#         folder = service.files().get(
#             fileId=MASTER_FOLDER_ID, 
#             supportsAllDrives=True
#         ).execute()
#         print(f"✅ SUCCESS! The Service Account can see the folder. Its name is: {folder.get('name')}")
#     except Exception as e:
#         print(f"❌ PERMISSION DENIED: The Service Account cannot see the folder at all.\n{e}")

# --- RECURSIVE FOLDER SCANNER ---
def walk_drive_folder(service, folder_id, current_local_path, current_relative_path, plan_list):
    query = f"'{folder_id}' in parents and trashed = false"
    
    page_token = None
    while True:
        response = service.files().list(
            q=query, 
            spaces='drive', 
            fields='nextPageToken, files(id, name, mimeType, size, modifiedTime)', 
            pageToken=page_token,
            includeItemsFromAllDrives=True, 
            supportsAllDrives=True,
            corpora='allDrives'
        ).execute()

        for item in response.get('files', []):
            if item['mimeType'] == 'application/vnd.google-apps.folder':
                new_local_path = os.path.join(current_local_path, item['name'])
                new_relative_path = os.path.join(current_relative_path, item['name']) if current_relative_path else item['name']
                walk_drive_folder(service, item['id'], new_local_path, new_relative_path, plan_list)
            else:
                # --- NEW: SYNC STATUS COMPARISON ---
                expected_local_path = os.path.join(current_local_path, item['name'])
                sync_status = "missing" # Default state
                
                drive_time_str = item.get('modifiedTime')
                
                if os.path.exists(expected_local_path) and drive_time_str:
                    try:
                        import datetime # Placed here to guarantee it grabs the main module
                        
                        # 1. Clean the Google Drive string (strips off milliseconds and 'Z')
                        # "2023-10-24T14:30:00.000Z" -> "2023-10-24T14:30:00"
                        clean_time = drive_time_str.split('.')[0].replace('Z', '')
                        
                        # 2. Parse it safely into a standard UNIX timestamp
                        drive_dt = datetime.datetime.strptime(clean_time, "%Y-%m-%dT%H:%M:%S")
                        drive_ts = drive_dt.timestamp()
                        
                        # 3. Get the Mac's local modified timestamp
                        local_ts = os.path.getmtime(expected_local_path)
                        
                        # 4. Compare! (Adding a 2-second buffer for OS quirks)
                        if drive_ts > (local_ts + 2):
                            sync_status = "outdated" # Drive is newer than Mac
                        else:
                            sync_status = "up-to-date" # Mac is identical or newer
                            
                    except Exception as e:
                        print(f"Time parse error for {item['name']}: {e}")
                # -----------------------------------

                plan_list.append({
                    "id": item['id'],
                    "name": item['name'],
                    "path": current_local_path, 
                    "relative_folder": current_relative_path or "Root Folder",
                    "size": int(item.get('size', 0)),
                    "modifiedTime": drive_time_str,
                    "sync_status": sync_status # <--- Send to Electron!
                })

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break

# ==========================================
# 🚀 ROUTE 1: SCAN & PREVIEW
# ==========================================
@app.route('/local/drive-scan', methods=['POST']) # CHANGED FROM GET TO POST
def drive_scan():
    try:
        # Get custom path from frontend, fallback to LOCAL_BACKUP_DIR
        data = request.json or {}
        custom_path = data.get('backup_path', LOCAL_BACKUP_DIR)
        
        service = get_drive_service() 
        plan_list = []
        
        print(f"🔍 Scanning Drive Folder: {MASTER_FOLDER_ID}")
        print(f"💾 Target Local Path: {custom_path}")
        
        # Pass the custom_path into the scanner instead of the hardcoded one
        walk_drive_folder(service, MASTER_FOLDER_ID, custom_path, "", plan_list)
        
        with open(PLAN_FILE, 'w') as f:
            json.dump(plan_list, f, indent=4)
            
        total_size_mb = sum([f['size'] for f in plan_list]) / (1024 * 1024)
            
        return jsonify({
            "success": True, 
            "total_files": len(plan_list),
            "total_size_mb": round(total_size_mb, 2),
            "files": plan_list,
            "used_path": custom_path # Send back the path we actually used
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/local/validate-path', methods=['POST'])
def validate_path():
    data = request.json or {}
    new_path = data.get('new_path', '').strip()
    current_path = data.get('current_path', '').strip()

    # --- NEW: Helper function to ignore invisible system files ---
    def get_real_contents(folder_path):
        if not os.path.exists(folder_path):
            return []
        all_files = os.listdir(folder_path)
        # Filter out Mac and Windows hidden system files
        return [f for f in all_files if f not in ['.DS_Store', 'Thumbs.db', 'desktop.ini']]

    # --- 1. VALIDATE THE CURRENT (SOURCE) FOLDER ---
    if current_path and current_path != new_path and os.path.exists(current_path):
        current_contents = get_real_contents(current_path)
        if len(current_contents) > 0:
            return jsonify({
                "valid": False, 
                "error": f"Your current backup folder is not empty. It contains {len(current_contents)} real item(s). Please move or delete them before switching."
            })

    if not new_path:
        return jsonify({"valid": True})

    # --- 2. VALIDATE THE NEW (TARGET) FOLDER ---
    if not os.path.isabs(new_path):
        return jsonify({"valid": False, "error": "New path must be an absolute path (e.g., /Users/... or C:\\...)"})

    if os.path.exists(new_path):
        new_contents = get_real_contents(new_path)
        if len(new_contents) > 0:
            return jsonify({
                "valid": False, 
                "error": f"The new target folder is not empty. It contains {len(new_contents)} real item(s). Please choose an empty folder."
            })

    return jsonify({"valid": True})

# ==========================================
# BACKGROUND DOWNLOAD ENGINE
# ==========================================
# 1. ADDED 'app' AND 'selected_ids' to arguments
def background_download_process(app, selected_ids):
    """Runs invisibly in the background so the API doesn't timeout."""
    
    # 2. ADDED APPLICATION CONTEXT BADGE
    with app.app_context():
        try:
            # 3. SWAPPED TO THE WORKING AUTHENTICATION
            service = get_drive_service()
            
            with open(PLAN_FILE, 'r') as f:
                full_plan_list = json.load(f)
                
            # 4. FILTER THE LIST: Only keep the files the user actually checked!
            plan_list = [f for f in full_plan_list if f['id'] in selected_ids]
            total_files = len(plan_list)

            if total_files == 0:
                with open(STATUS_FILE, 'w') as f:
                    json.dump({"state": "completed", "percent": 100, "message": "No new files needed to be downloaded."}, f)
                return
            
            for index, file_data in enumerate(plan_list):
                # Update the status file for Electron to read
                status = {
                    "state": "downloading",
                    "current_file": file_data['name'],
                    "progress": f"{index + 1} / {total_files}",
                    "percent": int(((index + 1) / total_files) * 100)
                }
                with open(STATUS_FILE, 'w') as f:
                    json.dump(status, f)

                # Create the exact folder structure on the Mac
                local_folder = file_data['path']
                os.makedirs(local_folder, exist_ok=True)
                local_file_path = os.path.join(local_folder, file_data['name'])

                # Download the file (Skip if it already exists!)
                if not os.path.exists(local_file_path):
                    print(f"📥 Downloading: {file_data['name']}")
                    request = service.files().get_media(fileId=file_data['id'])
                    fh = io.FileIO(local_file_path, 'wb')
                    downloader = MediaIoBaseDownload(fh, request)
                    
                    done = False
                    while done is False:
                        _, done = downloader.next_chunk()
                    
                    # Wait slightly so we don't spam Google's API limit
                    time.sleep(1) 
                else:
                    print(f"⏭️ Skipped (Already exists): {file_data['name']}")

            # Mark as finished!
            with open(STATUS_FILE, 'w') as f:
                json.dump({"state": "completed", "percent": 100, "message": "All selected files downloaded successfully!"}, f)

        except Exception as e:
            with open(STATUS_FILE, 'w') as f:
                json.dump({"state": "error", "error": str(e)}, f)


# ==========================================
# 🚀 ROUTE 2: TRIGGER DOWNLOAD
# ==========================================
@app.route('/local/drive-start', methods=['POST'])
def drive_start():
    # 5. GET THE CHECKED IDs FROM ELECTRON
    data = request.json or {}
    selected_ids = data.get('selected_ids', [])

    if not selected_ids:
        return jsonify({"success": False, "error": "No files selected for download."}), 400

    # Reset status file
    with open(STATUS_FILE, 'w') as f:
        json.dump({"state": "initializing", "percent": 0}, f)
        
    # Start the background thread, passing the 'app' and the 'selected_ids'
    thread = threading.Thread(target=background_download_process, args=(app, selected_ids))
    thread.daemon = True
    thread.start()
    
    return jsonify({"success": True, "message": "Download engine started in background."})


# ==========================================
# 🚀 ROUTE 3: CHECK STATUS (For UI Polling)
# ==========================================
@app.route('/local/drive-status', methods=['GET'])
def drive_status():
    if not os.path.exists(STATUS_FILE):
        return jsonify({"state": "idle"})
        
    with open(STATUS_FILE, 'r') as f:
        status = json.load(f)
        
    return jsonify(status)


############################################################################
##############GOOGLE DRIVE FOLDER SCANNER & BACKUP CODE (ENDS)##############
############################################################################

@app.route('/hello', methods=['GET'])
def hello_world():
    return jsonify({
        "status": "online",
        "message": "Hello World! Your local server is working.",
        "timestamp": time.strftime('%H:%M:%S')
    })


@app.route('/fix-stamp-one', methods=['POST'])
def fix_stamp_one():
    try:
        data = request.json['image']
        processed_img = remove_black_background_one(data)
        is_success, buffer = cv2.imencode(".png", processed_img)
        img_str = base64.b64encode(buffer).decode()
        return jsonify({"success": True, "image": f"data:image/png;base64,{img_str}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/fix-stamp-two', methods=['POST'])
def fix_stamp_two():
    try:
        data = request.json['image']
        processed_img = remove_black_background_two(data)

        is_success, buffer = cv2.imencode(".png", processed_img)
        img_str = base64.b64encode(buffer).decode()

        return jsonify({
            "success": True,
            "image": f"data:image/png;base64,{img_str}"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/fix-stamp-three', methods=['POST'])
def fix_stamp_three():
    try:
        data = request.json['image']
        processed_img = remove_black_background_three(data)

        is_success, buffer = cv2.imencode(".png", processed_img)
        img_str = base64.b64encode(buffer).decode()

        return jsonify({
            "success": True,
            "image": f"data:image/png;base64,{img_str}"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/fix-stamp-four', methods=['POST'])
def fix_stamp_four():
    try:
        data = request.json['image']
        processed_img = remove_black_background_four(data)

        is_success, buffer = cv2.imencode(".png", processed_img)
        img_str = base64.b64encode(buffer).decode()

        return jsonify({
            "success": True,
            "image": f"data:image/png;base64,{img_str}"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/fix-stamp-five', methods=['POST'])
def fix_stamp_five():
    try:
        data = request.json['image']
        processed_img = remove_black_background_five(data)

        is_success, buffer = cv2.imencode(".png", processed_img)
        img_str = base64.b64encode(buffer).decode()

        return jsonify({
            "success": True,
            "image": f"data:image/png;base64,{img_str}"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/del-stamp-sel', methods=['POST'])
def del_stamp_sel():
    try:
        data = request.json
        image_b64 = data['image'].split(',')[1]
        mask_data = data['mask']  # This is the array of 0s and 1s from JS

        # 1. Decode Image
        nparr = np.frombuffer(base64.b64decode(image_b64), np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)
        h, w = img.shape[:2]

        # Ensure we are working in BGR (3 channels) first
        if img.shape[2] == 4:
            img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

        # 2. Reconstruct Mask
        # Convert list of 0/1 to numpy array and reshape to image size
        mask = np.array(mask_data, dtype=np.uint8).reshape(h, w)

        # 3. Create Alpha Channel
        # 255 = Visible, 0 = Transparent
        # We want the 'selected' areas (1s) to be 0 (Transparent)
        alpha = np.ones((h, w), dtype=np.uint8) * 255
        alpha[mask > 0] = 0

        # 4. Smooth the edges of the cut
        # This prevents the "jagged" look where you deleted the background
        alpha = cv2.GaussianBlur(alpha, (3, 3), 0)

        # 5. Merge BGR + Alpha
        b, g, r = cv2.split(img)
        rgba = cv2.merge([b, g, r, alpha])

        # 6. Encode and Return
        _, buffer = cv2.imencode('.png', rgba)
        result_b64 = base64.b64encode(buffer).decode('utf-8')

        return jsonify({
            "success": True,
            "image": f"data:image/png;base64,{result_b64}"
        })
    except Exception as e:
        print(f"Error in fix-stamp-four: {str(e)}")
        return jsonify({"success": False, "error": str(e)})


@app.route('/straighten-stamp', methods=['POST'])
def straighten_endpoint():
    try:
        data = request.json['image']
        # Convert Base64 back to OpenCV
        nparr = np.frombuffer(base64.b64decode(data.split(',')[1]), np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_UNCHANGED)

        # Apply the straightening
        straightened_img = straighten_stamp(img)

        # Encode back to Base64
        is_success, buffer = cv2.imencode(".png", straightened_img)
        img_str = base64.b64encode(buffer).decode()

        return jsonify({"success": True, "image": f"data:image/png;base64,{img_str}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/remove-cancellation', methods=['POST'])
def remove_cancellation():
    try:
        data = request.json['image']
        processed_img = remove_cancellation_ink(data)
        is_success, buffer = cv2.imencode(".png", processed_img)
        img_str = base64.b64encode(buffer).decode()
        return jsonify({"success": True, "image": f"data:image/png;base64,{img_str}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/upload-stamp-one', methods=['POST'])
def upload_stamp_one():
    try:
        data = request.json['image']  # base64 data URL

        public_url = upload_base64_to_freeimage(data)

        if public_url:
            return jsonify({
                "success": True,
                "url": public_url
            })
        else:
            return jsonify({
                "success": False,
                "error": "Failed to upload to catbox.moe"
            })

    except KeyError:
        return jsonify({
            "success": False,
            "error": "Missing 'image' field"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


@app.route('/local/process-remote-batch', methods=['GET'])
def process_remote_batch():
    try:

        CLOUD_URL = CLOUD_API_URL
        HEADERS = {'X-API-KEY': SECRET_API_KEY, 'Content-Type': 'application/json',
                   'ngrok-skip-browser-warning': 'true'}

        # 1. Get the list of 65 broken rows from Cloud
        # try:
        #     fetch_resp = requests.get(f"{CLOUD_URL}/get-broken-stamps", headers=HEADERS)
        #     broken_rows = fetch_resp.json().get('data', [])
        # except Exception as e:
        #     return jsonify({"error": f"Failed to connect to cloud: {e}"})

        # results = []
        # 1. CALL ONLINE API: Get the list of 100 images
        print(f"🌍 Fetching batch from: {CLOUD_URL}")
        response = requests.get(CLOUD_URL + "/api/get-stamps-batch", headers=HEADERS, timeout=30)

        if response.status_code != 200:
            return jsonify({"error": "Failed to connect to online API", "details": response.text}), 500

        batch_data = response.json()

        # Check if list is empty
        if not batch_data:
            print("⚠️ No stamps returned from server.")
            return jsonify({"message": "No stamps to process."})

        print(f"📥 Received {len(batch_data)} stamps. Starting processing...")

        results = []

        for stamp in batch_data:
            stamp_id = stamp.get('id')
            file_name = stamp.get('fileName')
            image_path = stamp.get('drive_url')

            if not image_path:
                print(f"❌ Skipping {file_name}: No image path.")
                continue

            # Construct full URL
            if image_path.startswith('http'):
                full_image_url = image_path
            else:
                print(f"❌ Skipping {file_name}: Invalid URL format.")
                continue

            try:
                # 3. DOWNLOAD IMAGE
                img_resp = requests.get(full_image_url, timeout=10)

                if img_resp.status_code == 200:
                    # 4. CALCULATE VECTOR
                    img = Image.open(BytesIO(img_resp.content))
                    vector = imgmodel.encode(img)  # Ensure 'imgmodel' is defined
                    vector_list = vector.tolist()
                    vector_json_string = json.dumps(vector_list)
                    # 5. CONSOLE IT
                    print(f"✅ Processed Id {id} - {file_name}")
                    print(f"✅ Processed {file_name}")
                    print(f"Vector preview: {vector_list[:5]}")

                    results.append({
                        "fileName": file_name,
                        "status": "Vector Calculated",
                        "vector_preview": vector_list[:5]
                    })

                    if stamp_id:
                        update_payload = {
                            "id": stamp_id,
                            "updates": {
                                "image_vector": vector_json_string
                            }
                        }

                        print(f"📤 Uploading vector for ID {stamp_id} ({file_name})...")

                        try:
                            # Make sure this URL is correct
                            update_url = f"{CLOUD_URL}/admin/update-stamp"

                            update_resp = requests.post(
                                update_url,
                                json=update_payload,
                                headers=HEADERS,
                                timeout=10
                            )

                            if update_resp.status_code == 200:
                                print(f"✅ Cloud Update Success for {file_name}!")
                            else:
                                print(f"⚠️ Update Failed: {update_resp.status_code} - {update_resp.text}")

                        except Exception as upload_err:
                            print(f"❌ Connection Error during upload: {upload_err}")
                    else:
                        print(f"⚠️ Cannot update {file_name}: Missing 'id'.")

                    # --- ADDED REST HERE ---
                    print("💤 Resting for 2 seconds...")
                    time.sleep(2)
                    # -----------------------

                else:
                    print(f"❌ Failed to download image for {file_name}")

            except Exception as inner_e:
                print(f"❌ Error processing {file_name}: {inner_e}")
                # Optional: You might want to rest even on error, or skip resting
                # time.sleep(10)

        return jsonify({
            "success": True,
            "processed": len(results),
            "data": results
        })

    except Exception as e:
        print(f"🔥 Critical Error: {e}")
        return jsonify({"error": str(e)}), 500


import threading
import json
import os

SYNC_STATUS_FILE = 'db_sync_status.json'


# 1. ADDED 'app' TO THE FUNCTION ARGUMENTS
def run_sync_background(app):
    # 2. ADDED THE APPLICATION CONTEXT BADGE
    with app.app_context():
        try:
            # Write initial status
            with open(SYNC_STATUS_FILE, 'w') as f:
                json.dump({"status": "running", "logs": "Starting mass repair and sync..."}, f)

            # ==========================================================
            # 🚨 YOUR EXISTING SYNC-DB LOGIC
            # ==========================================================

            try:
                # 1. Hardcode the exact path to your db_sync.py file!
                script_path = '/Users/stampseva/Public/dbsync/db_sync.py'
                working_dir = '/Users/stampseva/Public/dbsync'

                print(f"🚀 API Triggering database sync script at: {script_path}")

                # 2. Run the script and capture the terminal output
                result = subprocess.run(
                    ['python3', script_path],
                    cwd=working_dir,
                    capture_output=True,
                    text=True
                )

                # 3. Check if the script was successful
                if result.returncode == 0:
                    print("✅ Sync script finished successfully.")

                    # CHANGED: Write success to the file instead of returning jsonify
                    with open(SYNC_STATUS_FILE, 'w') as f:
                        json.dump({
                            "status": "completed",
                            "success": True,
                            "logs": result.stdout
                        }, f)

                else:
                    # If the script crashed
                    print(f"❌ Sync Script Error:\n{result.stderr}")

                    # CHANGED: Write error to the file instead of returning jsonify
                    with open(SYNC_STATUS_FILE, 'w') as f:
                        json.dump({
                            "status": "error",
                            "success": False,
                            "error": "The sync script encountered an error.\n" + result.stderr,
                            "logs": result.stdout
                        }, f)

            except Exception as e:
                print(f"🔥 Critical API Error: {e}")

                # CHANGED: Write critical error to file
                with open(SYNC_STATUS_FILE, 'w') as f:
                    json.dump({
                        "status": "error",
                        "success": False,
                        "error": str(e)
                    }, f)

        except Exception as e:
            # If the outer thread crashes entirely
            with open(SYNC_STATUS_FILE, 'w') as f:
                json.dump({"status": "error", "success": False, "error": str(e)}, f)



@app.route('/local/sync-db', methods=['POST'])
def trigger_sync():
    # Pass 'app' into args=(app,)
    thread = threading.Thread(target=run_sync_background, args=(app,))
    thread.daemon = True
    thread.start()
    return jsonify({"success": True, "message": "Sync started in background."})


# --- 3. THE POLLING ROUTE ---
@app.route('/local/sync-status', methods=['GET'])
def get_sync_status():
    if not os.path.exists(SYNC_STATUS_FILE):
        return jsonify({"status": "idle"})

    with open(SYNC_STATUS_FILE, 'r') as f:
        status = json.load(f)

    return jsonify(status)

# @app.route('/local/sync-db', methods=['POST', 'GET'])
# def api_sync_db():
#     try:
#         # 1. Hardcode the exact path to your db_sync.py file!
#         script_path = '/Users/stampseva/Public/dbsync/db_sync.py'
#         working_dir = '/Users/stampseva/Public/dbsync'  # Good practice to run it from its own folder
#
#         print(f"🚀 API Triggering database sync script at: {script_path}")
#
#         # 2. Run the script and capture the terminal output
#         # We add cwd=working_dir so Python runs it as if you were standing in that folder
#         result = subprocess.run(
#             ['python3', script_path],
#             cwd=working_dir,
#             capture_output=True,
#             text=True
#         )
#
#         # 3. Check if the script was successful
#         if result.returncode == 0:
#             print("✅ Sync script finished successfully.")
#             return jsonify({
#                 "success": True,
#                 "message": "Databases synced successfully!",
#                 "logs": result.stdout  # This sends all the print() statements to Electron!
#             })
#         else:
#             # If the script crashed (e.g. wrong password or tunnel failed)
#             print(f"❌ Sync Script Error:\n{result.stderr}")
#             return jsonify({
#                 "success": False,
#                 "error": "The sync script encountered an error.",
#                 "logs": result.stdout,  # Shows what printed before the crash
#                 "error_logs": result.stderr  # Shows the actual crash error
#             }), 500
#
#     except Exception as e:
#         print(f"🔥 Critical API Error: {e}")
#         return jsonify({"success": False, "error": str(e)}), 500


# @app.route('/search-stampone', methods=['POST'])
# def search_stamp():
#     try:
#         image_url = request.json['image_url']

#         result = run_reverse_image_search(
#             image_url=image_url,
#             api_key=SERPAPI_API_KEY
#         )

#         if result["success"]:
#             return jsonify({
#                 "success": True,
#                 "data": result["data"]
#             })

#         return jsonify({
#             "success": False,
#             "error": result["error"]
#         })

#     except KeyError:
#         return jsonify({
#             "success": False,
#             "error": "Missing 'image_url' field"
#         })

#     except Exception as e:
#         return jsonify({
#             "success": False,
#             "error": str(e)
#         })
# --- Minimal Endpoint ---
@app.route('/search-stamp', methods=['POST'])
def search_stamp():
    try:
        data = request.json
        # 1. Get raw details from your AI function
        result = get_stamp_details(data.get('image'))

        # 2. Extract the messy theme the AI returned
        # (Assuming 'theme' is a key in your result dictionary)
        # 1. Access the raw AI theme using your specific key 'THEME'
        raw_theme = result.get('THEME', 'Others')

        # 2. Map it to your fixed list
        clean_theme = get_mapped_theme(raw_theme)

        # 3. Update the result so it only contains your fixed category
        result['THEME'] = clean_theme

        # 4. Return the result with the cleaned theme
        return jsonify({"success": True, "data": result})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/save-to-libre', methods=['POST'])
def save_to_libre():
    try:
        data = request.json
        folder_name = data.get('folder', 'Stamp_Collection')
        image_path = data.get('imagePath')
        # We use .xlsx because LibreOffice opens it natively and it is much more stable for images
        file_path = f"{folder_name}.xlsx"

        # 1. Define the Exact Sequence you requested
        base_headers = [
            "File name", "Image", "Country", "Theme",
            "Printed price", "Year", "Extra copies",
            "Estimated value", "Colour", "Initials", "Remarks"
        ]

        # 2. Map incoming JSON keys to these Headers
        key_map = {
            'fileName': "File name", 'Country': "Country", 'THEME': "Theme",
            'printed_price': "Printed price", 'Year': "Year", 'extra_copies': "Extra copies",
            'estimated_Value': "Estimated value", 'Colour': "Colour", 'Initials': "Initials", 'Remarks': "Remarks"
        }

        # 3. Load or Create Workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Stamps"
            # Write initial headers
            for col, header in enumerate(base_headers, 1):
                ws.cell(row=1, column=col).value = header

        # 4. Identify all headers currently in the sheet (to handle dynamic AI fields)
        current_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        # 5. Prepare data row
        row_idx = ws.max_row + 1
        ws.row_dimensions[row_idx].height = 80  # Make row tall for the image

        # Process all data sent by Gemini
        for raw_key, value in data.items():
            if raw_key in ['folder', 'imagePath', 'timestamp']: continue

            header_name = key_map.get(raw_key, raw_key.replace('_', ' ').title())

            # If header doesn't exist, add it to the top row
            if header_name not in current_headers:
                new_col = len(current_headers) + 1
                ws.cell(row=1, column=new_col).value = header_name
                current_headers.append(header_name)

            # Find the column index for this header and write the value
            col_idx = current_headers.index(header_name) + 1
            if header_name != "Image":  # Skip image for text writing
                ws.cell(row=row_idx, column=col_idx).value = str(value)

        # 6. Insert the Actual Image into Column 2 ("Image")
        if image_path and os.path.exists(image_path):
            # Resize image to thumbnail so file doesn't get too huge
            img = PILImage.open(image_path)
            img.thumbnail((100, 100))
            img_path_temp = f"temp_{os.path.basename(image_path)}"
            img.save(img_path_temp)

            xl_img = XLImage(img_path_temp)
            # Anchor to Column B (Index 2)
            ws.add_image(xl_img, f'B{row_idx}')

        # 7. Final Save
        wb.save(file_path)

        # Cleanup temp image
        if os.path.exists(img_path_temp):
            os.remove(img_path_temp)

        return jsonify({"success": True, "message": "Saved to Spreadsheet!"})

    except Exception as e:
        print(f"ERROR: {str(e)}")
        return jsonify({"success": False, "error": str(e)})


@app.route('/run-mass-repair', methods=['GET'])
def run_mass_repair():
    print("🚀 Starting Mass Hash Repair...")

    CLOUD_URL = CLOUD_API_URL
    HEADERS = {'X-API-KEY': SECRET_API_KEY, 'Content-Type': 'application/json', 'ngrok-skip-browser-warning': 'true'}

    # 1. Get the list of 65 broken rows from Cloud
    try:
        fetch_resp = requests.get(f"{CLOUD_URL}/get-broken-stamps", headers=HEADERS)
        broken_rows = fetch_resp.json().get('data', [])
    except Exception as e:
        return jsonify({"error": f"Failed to connect to cloud: {e}"})

    results = []

    # 2. Loop through each broken row
    for row in broken_rows:
        try:
            row_id = row['id']
            img_url = row['drive_url']

            print(f"🔄 Processing ID {row_id}: {row['fileName']}...")

            # 3. Download image into RAM
            img_data = requests.get(img_url, timeout=10).content
            img = Image.open(io.BytesIO(img_data))

            # 4. Generate local hashes
            payload = {
                "id": row_id,
                "phash": str(imagehash.phash(img)),
                "dhash": str(imagehash.dhash(img)),
                "whash": str(imagehash.whash(img))
            }

            # 5. Send hashes back to Cloud Repair endpoint
            # (Assumes you added the /repair-cloud-row endpoint from the previous prompt)
            update_resp = requests.post(f"{CLOUD_URL}/repair-cloud-row", json=payload, headers=HEADERS)

            results.append({
                "id": row_id,
                "file": row['fileName'],
                "status": "Success" if update_resp.status_code == 200 else "Failed"
            })

        except Exception as e:
            print(f"❌ Error on ID {row_id}: {e}")
            results.append({"id": row_id, "error": str(e)})

    return jsonify({
        "total_processed": len(results),
        "details": results
    })

# --- ROUTE 1: THE PREVIEW (Dry Run) ---
@app.route('/local/sync-preview', methods=['GET'])
def sync_preview():
    # 🚨 Grab the date from the frontend URL
    frontend_last_sync = request.args.get('last_sync')

    # === DEBUG PRINT START ===
    print("\n" + "=" * 50)
    print("🔍 SYNC PREVIEW TRIGGERED")
    print(f"📥 Received last_sync from frontend: '{frontend_last_sync}'")
    print("=" * 50)
    # === DEBUG PRINT END ===

    preview_report = []
    total_pending = 0

    # 🚨 NEW: Dictionary to hold the actual raw data to send to Electron
    detailed_records = {}

    try:
        local_conn = mysql.connector.connect(**db_config)
        local_cursor = local_conn.cursor(dictionary=True)
        local_cursor.execute("SET time_zone = '+00:00'")

        for table, columns in SYNC_BLUEPRINT.items():
            # Decide which date to use: Frontend > File > Fallback
            last_sync = "2026-03-19 03:00:00"

            if frontend_last_sync:
                last_sync = frontend_last_sync
            else:
                sync_file = f"last_sync_{table}.txt"
                if os.path.exists(sync_file):
                    with open(sync_file, 'r') as f:
                        last_sync = f.read().strip()

            columns_sql = ", ".join([f"`{col}`" for col in columns])
            if table == "duplicate_audit":
                sql = f"SELECT {columns_sql} FROM `{table}` WHERE checked_at > %s"
            else:
                sql = f"SELECT {columns_sql} FROM `{table}` WHERE updated_at > %s"

            # === DEBUG PRINT SQL ===
            print(f"\n⚙️ Analyzing Table: `{table}`")
            print(f"   📅 Target Date Used: {last_sync}")
            print(f"   📜 Executing SQL: {sql}")

            local_cursor.execute(sql, (last_sync,))
            records = local_cursor.fetchall()

            # === DEBUG PRINT RESULTS ===
            print(f"   ✅ Found {len(records)} records modified after {last_sync}.")

            if records:
                total_pending += len(records)
                preview_report.append(f"📦 {table}: {len(records)} records pending.")

                # 🚨 NEW: Make dates JSON-safe, then save to our detailed_records dictionary
                for record in records:
                    for key, value in record.items():
                        # If the value is a datetime object, convert it to a string
                        if hasattr(value, 'strftime'):
                            record[key] = value.strftime('%Y-%m-%d %H:%M:%S')

                detailed_records[table] = records

        local_cursor.close()
        local_conn.close()

        # === DEBUG PRINT END OF RUN ===
        print(f"\n🏁 PREVIEW COMPLETE: Total pending records = {total_pending}")
        print("=" * 50 + "\n")

        if total_pending == 0:
            return jsonify(
                {"success": True, "status": "empty", "message": f"Everything is up to date since {last_sync}!"})

        return jsonify({
            "success": True,
            "status": "pending",
            "total": total_pending,
            "report": "\n".join(preview_report),
            "used_date": last_sync,
            "details": detailed_records  # 🚨 NEW: Sending the actual data back!
        })

    except Exception as e:
        # === DEBUG PRINT ERROR ===
        print(f"\n❌ CRITICAL ERROR IN SYNC PREVIEW: {str(e)}")
        print("=" * 50 + "\n")
        return jsonify({"success": False, "error": str(e)})


@app.route('/local/sync-commit', methods=['POST'])
def sync_commit():
    data = request.json or {}
    frontend_last_sync = data.get('last_sync')
    is_dry_run = data.get('is_dry_run')

    print("\n" + "=" * 50)
    print("🚀 API SYNC COMMIT TRIGGERED")
    print("=" * 50)

    if not frontend_last_sync:
        return jsonify({"success": False, "error": "No sync date provided by the frontend!"}), 400

    try:
        print("🔌 Connecting to Local Database...")
        local_conn = mysql.connector.connect(**db_config)
        local_cursor = local_conn.cursor(dictionary=True)
        local_cursor.execute("SET time_zone = '+00:00'")

        total_pushed = 0
        master_logs = []
        current_sync_start = dt.datetime.now(dt.timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

        for table, columns in SYNC_BLUEPRINT.items():
            print(f"\n⚙️ Analyzing Table: `{table}`")

            # Fetch local changes
            columns_sql = ", ".join([f"`{col}`" for col in columns])
            if table == "duplicate_audit":
                fetch_sql = f"SELECT {columns_sql} FROM `{table}` WHERE checked_at > %s"
            else:
                fetch_sql = f"SELECT {columns_sql} FROM `{table}` WHERE updated_at > %s"

            local_cursor.execute(fetch_sql, (frontend_last_sync,))
            records_to_sync = local_cursor.fetchall()

            if not records_to_sync:
                print(f"   ⏭️ Up to date. Skipping.")
                continue

            # Format values and safely convert MySQL datetimes into JSON-friendly strings
            values = []
            for record in records_to_sync:
                row = []
                for col in columns:
                    val = record.get(col)

                    # If the value is a date or time object, convert it to a string!
                    if hasattr(val, 'strftime'):
                        val = val.strftime('%Y-%m-%d %H:%M:%S')

                    row.append(val)
                values.append(row)

            # 📦 Chunking Logic (25 records per request)
            chunk_size = 25
            table_pushed = 0

            for i in range(0, len(values), chunk_size):
                chunk = values[i: i + chunk_size]

                payload = {
                    "table": table,
                    "columns": columns,
                    "values": chunk
                }

                print(f"   📤 Sending chunk ({len(chunk)} records) to Cloud API...")

                print(f"   📤 Dry run value ( {is_dry_run} ) from electron API...")

                # 🚦 IF IT IS NOT A DRY RUN (is_dry_run is exactly the boolean False)
                if is_dry_run is False:
                    print("   🔥 [LIVE SYNC] Executing real push to cloud...")

                    # Send HTTP Request to Cloud
                    cloud_res = requests.post(
                        f"{CLOUD_API_URL}/api/cloud/receive-sync-chunk",
                        json=payload,
                        headers={'X-API-KEY': SECRET_API_KEY},
                        timeout=30
                    )

                    # 🚨 THE NEW SAFETY CATCH 🚨
                    try:
                        cloud_data = cloud_res.json()
                    except Exception as e:
                        # If it's not JSON, print exactly what the cloud server returned!
                        print(f"   ❌ FATAL: Cloud server did not return JSON!")
                        print(f"   HTTP Status: {cloud_res.status_code}")
                        print(
                            f"   Raw Response: {cloud_res.text[:500]}")  # Prints the first 500 characters of the error page
                        raise Exception(
                            f"Cloud API crashed! Status Code: {cloud_res.status_code}. Check Python terminal for details.")

                    if not cloud_data.get('success'):
                        raise Exception(f"Cloud Server Error: {cloud_data.get('error')}")

                    print("   ✅ Chunk accepted by Cloud.")

                # 🚦 IF IT IS A DRY RUN (is_dry_run is True)
                else:
                    print("   🧪 [DRY RUN] Skipping live cloud request. Simulation only.")

                table_pushed += len(chunk)
                print("   ✅ Chunk accepted by Cloud.")

                # ⏳ Wait 20 seconds IF this is not the last chunk for this table
                if i + chunk_size < len(values):
                    print("   ⏳ Waiting 20 seconds to prevent rate limiting...")
                    time.sleep(20)

            # =================================================================
            # 🚨 INDENTATION FIXED BELOW: Moved OUT of the 'if' block & chunk loop
            # =================================================================

            # Update Tracker File
            sync_file = f"last_sync_{table}.txt"
            with open(sync_file, 'w') as f:
                f.write(current_sync_start)

            total_pushed += table_pushed

            # ==========================================
            # 📝 THE NEW ADVANCED LOGGING SYSTEM
            # ==========================================

            # 1. Figure out the Primary Key for the verification query
            pk_col = 'id'
            if 'stamp_id' in columns:
                pk_col = 'stamp_id'
            elif pk_col not in columns:
                pk_col = columns[0]  # Fallback to first column

            # 2. Grab the IDs of the records we just pushed
            pushed_ids = [str(record.get(pk_col)) for record in records_to_sync]

            # 3. Create the Verification SQL Query (Cap at 20 IDs so the UI doesn't freeze on huge syncs)
            display_ids = pushed_ids[:20]
            in_clause = ", ".join([f"'{pk}'" for pk in display_ids])
            verify_sql = f"SELECT * FROM `{table}` WHERE `{pk_col}` IN ({in_clause});"
            if len(pushed_ids) > 20:
                verify_sql += f" -- (Query limited to first 20 IDs out of {len(pushed_ids)})"

            # 4. Create a readable preview of the Data Payload
            data_preview = ""
            for row in values[:5]:  # Show max 5 rows of data so the terminal is readable
                data_preview += str(row) + "\n"
            if len(values) > 5:
                data_preview += f"... and {len(values) - 5} more records."

            # 5. Build the beautiful UI Log Block
            detailed_log = f"""
==================================================
📦 TABLE: {table} ({len(records_to_sync)} records synced)
==================================================
📤 DATA SENT TO CLOUD (Preview):
{data_preview}

🔍 VERIFICATION QUERY (Copy/Paste into Cloud MySQL):
{verify_sql}
"""
            master_logs.append(detailed_log)

        # ----------------------------------------------
        # End of the Table Loop
        # ----------------------------------------------

        # 🚨 INDENTATION FIXED BELOW: Moved properly INSIDE the try block
        local_cursor.close()
        local_conn.close()

        print(f"\n🏁 API SYNC COMPLETE! Total records: {total_pushed}")
        print("=" * 50 + "\n")

        return jsonify({
            "success": True,
            "logs": "\n".join(
                master_logs) + f"\n\n🎉 Sync Complete via API! Total records pushed: {total_pushed}"
        })

    except Exception as e:
        # 🚨 INDENTATION FIXED BELOW: Moved properly INSIDE the except block
        print(f"\n❌ CRITICAL ERROR IN SYNC COMMIT: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/check-duplicate', methods=['POST'])
def check_duplicate():
    print("\n--- Starting High-Scale Cloud Duplicate Check (Normalized) ---")

    # searchMethod
    # <option value="refined">✨ High-Precision Refined (Full Verification)</option>
    # <option value="combined">🔗 Combined Search (AI Vectors + Hashes)</option>
    # <option value="faiss_only">🧠 Visual AI Only (FAISS Vector Search)</option>
    # <option value="hashes_only">🧩 Fingerprint Only (Image Hashes)</option>

    try:
        # Step 1: Process Local Image Data
        data = request.json
        if not data or 'image' not in data:
            print("❌ Error: No image data received from frontend")
            return jsonify({"success": False, "error": "No image data"}), 400

        incoming_filename = data.get('fileName')
        country = data.get('country')
        searchMethod = data.get('searchMethod', 'refined')  # Default to 'refined' if not provided

        # Now you can safely check if you actually got a country or not
        if country is None:
            print("No country was provided in the payload!")
            # You can set a default here, or just ignore it
            country = ""
        else:
            print(f"We received the country: {country}")

        # Clean the base64 string
        image_raw = data.get('image')
        image_b64 = image_raw.split(",")[1] if "," in image_raw else image_raw

        print(f"\n--- Search method came to python '{searchMethod}' ---")

        # --- NEW CODE: CALL FAISS ENDPOINT & PRINT RESULTS ---
        print("🔄 Calling local /search-faiss endpoint to find similar images...")
        try:
            # Dynamically build the URL to point to your own app
            faiss_url = f"{request.host_url.rstrip('/')}/search-faiss"

            # 1. Define the headers needed to bypass ngrok
            headers = {
                "ngrok-skip-browser-warning": "true"
            }

            # 2. Send the request to the FAISS endpoint with the headers included
            faiss_response = requests.post(
                faiss_url,
                json={"image": image_b64, "country": country},
                headers=headers,  # <--- ADD THIS LINE
                timeout=30
            )

            # Parse the JSON response
            faiss_result = faiss_response.json()

            # Print the formatted results to the terminal
            print("\n📊 --- faiss_result SEARCH RESULTS ---")
            print(f"🔍 Total faiss_result : {len(faiss_result.get('candidates', []))}")
            print("------------------------------\n")

        except requests.exceptions.RequestException as e:
            print(f"❌ Failed to reach local /search-faiss endpoint: {e}")
        # -----------------------------------------------------

        #################----------------------------------------
        print(f"\n--- if Search method is faiss_only then return result from here '{searchMethod}' ---")
        if searchMethod == "faiss_only":
            return jsonify({
                "success": True,
                "is_duplicate": len(faiss_result.get('candidates', [])) > 0,
                # True if it found anything, False if empty
                "candidates": faiss_result.get('candidates', [])
            })

        # --- NEW: Standardize the image before hashing ---
        img = get_standardized_img(image_b64)

        # Generate Hashes Locally
        current_phash = str(imagehash.phash(img))
        current_dhash = str(imagehash.dhash(img))
        current_whash = str(imagehash.whash(img))

        # Generate Vector Locally
        vector = imgmodel.encode(img)
        vector_list = vector.tolist()

        print(f"Vector preview: {vector_list[:5]}")
        print(
            f"✅ Normalized Hashes - pHash: {current_phash}, dHash: {current_dhash}, wHash: {current_whash} , incoming_filename: {incoming_filename}")

        # Step 2: Call Cloud Search
        cloud_url = f"{CLOUD_API_URL}/check-cloud-duplicate"
        headers = {'X-API-KEY': SECRET_API_KEY, 'Content-Type': 'application/json',
                   'ngrok-skip-browser-warning': 'true'}

        # Send Data (Including fileName for Priority 1 check)
        payload = {
            "phash": current_phash,
            "dhash": current_dhash,
            "whash": current_whash,
            "fileName": incoming_filename
        }

        print(f"🌐 Querying Cloud Database (Search Mode)...  {cloud_url} ")
        response = requests.post(cloud_url, json=payload, headers=headers, timeout=15)

        if response.status_code != 200:
            print(f"❌ Cloud API Error: {response.text}")
            return jsonify({"success": False, "error": f"Cloud API returned {response.status_code}"}), 500

        cloud_result = response.json()

        # print(f"🎯 CLOUD RESULT DUMP JSON: {json.dumps(cloud_result, indent=2)}")
        # ---------------------------------------------------------
        # PRIORITY 1: EXACT FILENAME MATCH (From Cloud)
        # ---------------------------------------------------------
        # If cloud says score is 100, it found the file by name. Return immediately.
        if cloud_result.get('score') == 100:
            print(f"🎯 EXACT FILENAME MATCH FOUND: {cloud_result.get('match_id')}")

            return jsonify({
                "success": True,
                "is_duplicate": True,
                "match_id": cloud_result.get('match_id'),
                "score": 100,
                "stamp_info": cloud_result.get('stamp_info'),
                "method": "Exact Filename Match"
            })

        # ---------------------------------------------------------
        # PRIORITY 2: DEEP MULTI-FUNCTION VERIFICATION
        # ---------------------------------------------------------
        candidates = cloud_result.get('candidates', [])
        print(f"🔍 Total Candidates after Cloud: {len(candidates)}")

        #################----------------------------------------
        print(f"\n--- if Search method is hashes_only then return result from here '{searchMethod}' ---")
        if searchMethod == "hashes_only":

            # --- NEW CODE: Print the scores to the console ---
            print("\n🧩 Printing Hashes Only Scores:")
            for i, candidate in enumerate(candidates):
                # Using .get() is safer just in case a candidate is missing the score key
                score = candidate.get('score', 'No score found')
                print(f"  -> Candidate {i + 1}: Score = {score}")
            # -------------------------------------------------

            return jsonify({
                "success": True,
                "is_duplicate": len(candidates) > 0,  # True if it found anything, False if empty
                "candidates": candidates
            })

        # 2. Merge in the FAISS candidates
        if 'faiss_result' in locals() and faiss_result.get('success'):
            faiss_candidates = faiss_result.get('candidates', [])

            # Create a fast-lookup set of IDs we already have from the Cloud
            # (You can use 'fileName' instead of 'id' here if you prefer)
            existing_ids = {c.get('id') for c in candidates}

            for fc in faiss_candidates:
                # Only add the FAISS match if the Cloud didn't already find it
                if fc.get('id') not in existing_ids:
                    candidates.append(fc)

                    # Add this ID to our tracking set just in case FAISS has internal duplicates
                    existing_ids.add(fc.get('id'))

        # 3. Optional: Sort the final combined list by score (highest to lowest)
        candidates = sorted(candidates, key=lambda x: x.get('score', 0), reverse=True)

        print(f"🔍 Total Candidates after merging Cloud and FAISS: {len(candidates)}")

        #################----------------------------------------
        print(f"\n--- if Search method is combined then return result from here '{searchMethod}' ---")

        if searchMethod == "combined":
            return jsonify({
                "success": True,
                "is_duplicate": len(candidates) > 0,  # True if it found anything, False if empty
                "candidates": candidates
            })

        if str(incoming_filename).strip().lower() == "noneall":
            print(f"\n📊 'NONEALL' TRIGGERED: Returning all {len(candidates)} merged matches.")

            return jsonify({
                "success": True,
                "is_duplicate": len(candidates) > 0,  # True if it found anything, False if empty
                "candidates": candidates
            })

        FINAL_THRESHOLD = 0.70  # Lower threshold to allow deep check
        qualifying_matches = []

        if candidates:
            print(f"\n🧠 AI Judging total no candidates: {len(candidates)} candidates from Cloud...")

            for cand in candidates:
                # --- A. GET AND CLEAN VECTOR DATA ---
                db_vector_data = cand.get('image_vector')
                votes = cand.get('votes', 0)
                if not db_vector_data: continue

                # 1. Parse JSON if it's a string
                if isinstance(db_vector_data, str):
                    try:
                        db_vector = json.loads(db_vector_data)
                    except:
                        print(f"     ⚠️ JSON Error for {cand.get('fileName')}")
                        continue
                else:
                    db_vector = db_vector_data

                # 2. SAFETY CHECK: Ensure db_vector is a list of numbers
                try:
                    if isinstance(db_vector, list) and len(db_vector) > 0 and isinstance(db_vector[0], dict):
                        print(f"     ❌ SKIP: Data corruption (Found dict inside vector) for {cand.get('fileName')}")
                        continue
                except:
                    pass

                # --- B. CALCULATE VECTOR SCORE ---
                try:
                    vector_score = util.cos_sim(vector_list, db_vector).item()
                except Exception as math_err:
                    print(f"     ❌ Math Error for {cand.get('fileName')}: {math_err}")
                    continue

                # --- C. QUALIFY & DEEP CHECK ---
                if vector_score >= FINAL_THRESHOLD:
                    print(f"   🔹 Candidate {cand['fileName']} qualified (Vector: {vector_score:.4f})")

                    # Prepare Remote Image
                    remote_url = cand.get('stamp_info', {}).get('drive_url')
                    remote_b64 = None

                    if remote_url:
                        try:
                            import base64
                            r_img = requests.get(remote_url, timeout=5)
                            if r_img.status_code == 200:
                                remote_b64 = base64.b64encode(r_img.content).decode('utf-8')
                        except:
                            pass

                    # Run Comparison Functions (AI + Feature Match)
                    ai_score_val = 0.0
                    feature_score_val = 0.0

                    if remote_b64:
                        try:
                            # Ensure we pass floats back to avoid 'dtype of dict' error
                            ai_score_val = float(compare_images_ai(image_b64, remote_b64, imgmodel))
                            feature_score_val = float(compare_images_features(image_b64, remote_b64))
                            print(f"     🧩 Feature Score: {feature_score_val:.2f} | 🤖 AI Score: {ai_score_val:.2f}")
                        except Exception as e:
                            print(f"     ⚠️ Comparison function error: {e}")

                    # Add to list (Ensure all values are simple types)
                    qualifying_matches.append({
                        "candidate": cand,
                        "votes": votes,
                        "vector_score": float(vector_score * 100),
                        "ai_score": ai_score_val,
                        "feature_score": feature_score_val
                    })

        #################----------------------------------------
        print(f"\n--- if Search method is combined then return result from here '{searchMethod}' ---")

        # ---------------------------------------------------------
        # FINAL VERDICT
        # ---------------------------------------------------------

        if qualifying_matches:
            # Sort safely using explicit float conversion
            # We sort by Vector Score (Similarity) to find the truest duplicate
            # Sort Priority: 1. AI Score -> 2. Vector Score -> 3. Feature Score
            # 1. Sort all the qualifying matches exactly as before (highest votes and scores first)
            qualifying_matches.sort(key=lambda x: (
                int(x['votes']),
                float(x['ai_score']),
                float(x['vector_score']),
                float(x['feature_score'])
            ), reverse=True)

            # 2. DECISION MAKING: Did the frontend send "None" as the filename?
            if str(incoming_filename).strip().lower() == "none":
                print(f"\n📊 BROWSER PREVIEW MODE: Returning all {len(qualifying_matches)} sorted matches.")

                # Format the entire array to look exactly like the single JSON response
                formatted_candidates = []
                for match in qualifying_matches:
                    cand_info = match['candidate']
                    formatted_candidates.append({
                        "is_duplicate": True,
                        "match_id": cand_info['fileName'],
                        "score": int(match['ai_score']),
                        "feature_score": int(match['feature_score']),
                        "stamp_info": cand_info.get('stamp_info', cand_info),  # Safe fallback
                        "method": "Multi-Function Verification",
                        "votes": match['votes']
                    })

                return jsonify({
                    "success": True,
                    "is_duplicate": True,
                    "candidates": formatted_candidates  # Sends the array back to the UI
                })

            # 3. NORMAL MODE: Just return the top winner
            else:
                winner = qualifying_matches[0]
                best_match_info = winner['candidate']

                print(f"\n🏆 WINNER CONFIRMED: {best_match_info['fileName']}")
                print(f"\n🏆 VOTES: {winner['votes']}")
                print(f"\n🏆 AI SCORE: {winner['ai_score']}")
                print(f"\n🏆 VECTOR SCORE: {winner['vector_score']}")
                print(f"\n🏆 FEATURE SCORE: {winner['feature_score']}")

                return jsonify({
                    "success": True,
                    "is_duplicate": True,
                    "match_id": best_match_info['fileName'],
                    "score": int(winner['ai_score']),
                    "feature_score": int(winner['feature_score']),
                    "stamp_info": best_match_info.get('stamp_info'),
                    "method": "Multi-Function Verification"
                })

        else:
            print(f"\n📉 No matches passed the {FINAL_THRESHOLD} threshold.")
            return jsonify({
                "success": True,
                "is_duplicate": False,
                "score": 0,
                "message": "No visual match found"
            })

    except Exception as e:
        import traceback
        print("🚨 SYSTEM CRASH IN LOCAL API:")
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/check-duplicate-app', methods=['POST'])
def check_duplicate_app():
    print("\n--- Starting Standardized Duplicate Check ---")
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({"success": False, "error": "No image data"}), 400

        # Clean the base64 string if it contains the data:image/jpeg prefix
        image_data = data.get('image')
        if "," in image_data:
            image_data = image_data.split(",")[1]

        image_raw = data.get('image')

        # --- NEW: DETECT IMAGE TYPE ---
        ext = "unknown"
        if "data:image/" in image_raw:
            # Extracts 'png', 'jpeg', or 'gif' from "data:image/png;base64,..."
            ext = image_raw.split(';')[0].split('/')[1]
            print(f"📸 Detected image type from header: .{ext}")

        # Clean the base64 string
        image_data = image_raw.split(",")[1] if "," in image_raw else image_raw

        # Fallback: If no header was present, check the binary signature
        if ext == "unknown":
            import base64
            header_bytes = base64.b64decode(image_data[:32])
            if header_bytes.startswith(b'\xff\xd8'):
                ext = "jpg"
            elif header_bytes.startswith(b'\x89PNG'):
                ext = "png"
            elif header_bytes.startswith(b'GIF8'):
                ext = "gif"
            print(f"🔍 Detected image type from binary: .{ext}")

        # --- NEW CODE: CALL FAISS ENDPOINT & PRINT RESULTS ---
        print("🔄 Calling local /search-faiss endpoint to find similar images...")
        try:
            # Dynamically build the URL to point to your own app
            faiss_url = f"{request.host_url.rstrip('/')}/search-faiss"

            # Send the request to the FAISS endpoint
            faiss_response = requests.post(
                faiss_url,
                json={"image": image_data},
                timeout=30
            )

            # Parse the JSON response
            faiss_result = faiss_response.json()

            # Print the formatted results to the terminal
            print("\n📊 --- FAISS SEARCH RESULTS ---")
            # print(json.dumps(faiss_result, indent=4))
            print("------------------------------\n")

        except requests.exceptions.RequestException as e:
            print(f"❌ Failed to reach local /search-faiss endpoint: {e}")
        # -----------------------------------------------------

        # --- NEW STANDARDIZATION STEP ---
        img = get_standardized_img(image_data)

        # Generate Hashes from the Standardized Image
        current_phash = str(imagehash.phash(img))
        current_dhash = str(imagehash.dhash(img))
        # --- ADDED: Calculate wHash ---
        current_whash = str(imagehash.whash(img))

        vector = imgmodel.encode(img)  # Ensure 'imgmodel' is defined
        vector_list = vector.tolist()

        print(f"Vector preview: {vector_list[:5]}")

        print(f"✅ Standardized Hashes - pHash: {current_phash}, dHash: {current_dhash}, wHash: {current_whash}")

        # Step 2: Call Cloud Search
        cloud_url = f"{CLOUD_API_URL}/check-cloud-duplicate-app"
        headers = {'X-API-KEY': SECRET_API_KEY, 'Content-Type': 'application/json',
                   'ngrok-skip-browser-warning': 'true'}

        # Payload now includes the trinity of hashes
        payload = {
            "phash": current_phash,
            "dhash": current_dhash,
            "whash": current_whash,
        }

        response = requests.post(cloud_url, json=payload, headers=headers, timeout=15)
        if response.status_code != 200:
            print(f"❌ Cloud API Error: {response.text}")
            print(f"❌ Cloud API Error: {response}")
            return jsonify({"success": False, "error": f"Cloud API returned {response.status_code}"}), 500

        # Step 3: Return Cloud Result directly to Electron
        # Step 3: Return Cloud Result directly to Electron
        cloud_result = response.json()
        candidates = cloud_result.get('candidates', [])
        esql = cloud_result.get('esql', 'N/A')
        # ---------------------------------------------------------
        # --- NEW: MULTI-FUNCTION VERIFICATION (The "Deep Check") ---
        # ---------------------------------------------------------

        # 2. Merge in the FAISS candidates
        if 'faiss_result' in locals() and faiss_result.get('success'):
            faiss_candidates = faiss_result.get('candidates', [])

            # Create a fast-lookup set of IDs we already have from the Cloud
            # (You can use 'fileName' instead of 'id' here if you prefer)
            existing_ids = {c.get('id') for c in candidates}

            for fc in faiss_candidates:
                # Only add the FAISS match if the Cloud didn't already find it
                if fc.get('id') not in existing_ids:
                    candidates.append(fc)

                    # Add this ID to our tracking set just in case FAISS has internal duplicates
                    existing_ids.add(fc.get('id'))

        # 3. Optional: Sort the final combined list by score (highest to lowest)
        candidates = sorted(candidates, key=lambda x: x.get('score', 0), reverse=True)

        # 1. Lower Threshold as requested
        # ---------------------------------------------------------
        # --- NEW: MULTI-FUNCTION VERIFICATION (With Safety Checks) ---
        # ---------------------------------------------------------
        print(f"\n🧠 AI SQL {esql}  from Cloud ...")
        FINAL_THRESHOLD = 0.75
        qualifying_matches = []

        if candidates:

            for cand in candidates:

                # --- A. GET AND CLEAN VECTOR DATA ---
                db_vector_data = cand.get('image_vector')
                hashes = cand.get('debug', {})

                # 2. Extract the distances (fallback to 99 if missing so math doesn't break)
                dist_p = hashes.get('dist_p', 99)
                dist_d = hashes.get('dist_d', 99)
                dist_w = hashes.get('dist_w', 99)
                votes = cand.get('votes', 0)

                print(
                    f"hash preview: {dist_p}, D hash preview: {dist_d}, W hash preview: {dist_w}, File Name: {cand.get('fileName')}")

                if not db_vector_data: continue

                # 1. Parse JSON if it's a string
                if isinstance(db_vector_data, str):
                    try:
                        db_vector = json.loads(db_vector_data)
                    except:
                        print(f"     ⚠️ JSON Error for {cand.get('fileName')}")
                        continue
                else:
                    db_vector = db_vector_data

                # 2. SAFETY CHECK: Ensure db_vector is a list of numbers
                # If the DB returns a list of dicts by accident, this prevents the crash.
                try:
                    if isinstance(db_vector, list) and len(db_vector) > 0 and isinstance(db_vector[0], dict):
                        print(f"     ❌ SKIP: Data corruption (Found dict inside vector) for {cand.get('fileName')}")
                        continue
                except:
                    pass

                # --- B. CALCULATE VECTOR SCORE ---
                try:
                    # util.cos_sim causes the "dtype of dict" error if inputs are wrong
                    vector_score = util.cos_sim(vector_list, db_vector).item()
                except Exception as math_err:
                    print(f"     ❌ Math Error for {cand.get('fileName')}: {math_err}")
                    continue

                # --- C. QUALIFY & DEEP CHECK ---
                if vector_score >= FINAL_THRESHOLD:
                    print(f"   🔹 Candidate {cand['fileName']} qualified (Vector: {vector_score:.4f})")

                    # Prepare Images
                    remote_url = cand.get('stamp_info', {}).get('drive_url')
                    remote_b64 = None

                    if remote_url:
                        try:
                            import base64
                            r_img = requests.get(remote_url, timeout=5)
                            if r_img.status_code == 200:
                                remote_b64 = base64.b64encode(r_img.content).decode('utf-8')
                        except:
                            pass

                    # Run Comparison Functions
                    ai_score_val = 0.0
                    feature_score_val = 0.0

                    if remote_b64:
                        try:
                            # Ensure we pass floats back
                            ai_score_val = float(compare_images_ai(image_data, remote_b64, imgmodel))
                            feature_score_val = float(compare_images_features(image_data, remote_b64))
                            print(f"     🧩 Feature Score: {feature_score_val:.2f} | 🤖 AI Score: {ai_score_val:.2f}")
                        except Exception as e:
                            print(f"     ⚠️ Comparison function error: {e}")

                    # Add to list (Ensure all values are simple types)
                    qualifying_matches.append({
                        "candidate": cand,
                        "votes": votes,
                        "vector_score": float(vector_score * 100),
                        "ai_score": ai_score_val,
                        "feature_score": feature_score_val
                    })

        # ---------------------------------------------------------
        # --- STEP 4: FINAL DECISION ---
        # ---------------------------------------------------------

        if qualifying_matches:
            # Sort safely using explicit float conversion
            # This prevents sorting errors if data is missing
            qualifying_matches.sort(key=lambda x: (
                int(x['votes']),
                float(x['ai_score']),
                float(x['vector_score']),
                float(x['feature_score'])
            ), reverse=True)
            print(f"\n🏆 NO OF QUALIFIED CONFIRMED: {len(qualifying_matches)}")
            winner = qualifying_matches[0]
            best_match_info = winner['candidate']

            print(f"\n🏆 WINNER CONFIRMED: {best_match_info['fileName']}")
            print(f"\n🏆 VOTES: {winner['votes']}")
            print(f"\n🏆 AI SCORE: {winner['ai_score']}")
            print(f"\n🏆 VECTOR SCORE: {winner['vector_score']}")
            print(f"\n🏆 FEATURE SCORE: {winner['feature_score']}")

            return jsonify({
                "success": True,
                "is_duplicate": True,
                "match_id": best_match_info['fileName'],
                "score": int(winner['ai_score']),
                "feature_score": int(winner['feature_score']),
                "stamp_info": best_match_info.get('stamp_info'),
                "method": "Multi-Function Verification"
            })

        else:
            print(f"\n📉 No matches passed the {FINAL_THRESHOLD} threshold.")
            return jsonify({
                "success": True,
                "is_duplicate": False,
                "score": 0,
                "message": "No visual match found"
            })

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/translate', methods=['POST'])
def translate_text():
    data = request.json
    text_to_translate = data.get('text')
    target = data.get('language')

    # --- DEBUGGING LOGS ---
    print(f"\n--- Translation Request ---")
    print(f"Incoming Target Language: '{target}'")
    print(f"Incoming Text (First 30 chars): {text_to_translate[:30]}...")

    # QUICK FIX: deep-translator usually needs 'zh-CN' for Mandarin
    if target == 'zh':
        print("Adjusting 'zh' to 'zh-CN' for GoogleTranslator compatibility.")
        target = 'zh-CN'

    try:
        translated = GoogleTranslator(source='auto', target=target).translate(text_to_translate)

        # --- DEBUGGING LOGS ---
        print(f"Translated Result (First 30 chars): {translated[:30]}...")

        if translated == text_to_translate:
            print("⚠️ WARNING: Output matches Input. Translation might have failed or target code is invalid.")

        return jsonify({"translated_text": translated})
    except Exception as e:
        print(f"❌ ERROR in translation: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/get-image-vector', methods=['POST'])
def get_image_vector():
    print("\n--- Generating Image Vector ---")
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({"success": False, "error": "No image data"}), 400

        # 1. Clean the Base64 String
        image_raw = data.get('image')
        image_data = image_raw

        # Remove header if present (e.g., "data:image/jpeg;base64,...")
        if "," in image_raw:
            image_data = image_raw.split(",")[1]

        # 2. Standardize Image (Resize/Convert)
        # using your existing helper function
        img = get_standardized_img(image_data)

        # 3. Generate Vector
        # model.encode returns a numpy array
        vector_numpy = imgmodel.encode(img)

        # 4. Convert to List
        vector_list = vector_numpy.tolist()

        vector_json_string = json.dumps(vector_list)

        print(f"✅ Vector Generated. Length: {len(vector_list)}")
        print(f"👀 Preview: {vector_list[:5]}...")

        return jsonify({
            "success": True,
            "vector": vector_json_string,
            "length": len(vector_list)
        })

    except Exception as e:
        print(f"❌ Error generating vector: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/start-audit', methods=['GET'])
def start_audit():
    global audit_running, audit_thread, run_count, start_time

    if audit_running:
        return jsonify({"message": "Auditor is already running!", "current_count": run_count})

    audit_running = True
    run_count = 0
    start_time = time.time()

    audit_thread = threading.Thread(target=periodic_auditor, daemon=True)
    audit_thread.start()

    return jsonify({
        "status": "Started",
        "message": "Auditor will hit the cloud every 2 minutes."
    })


@app.route('/stop-audit', methods=['GET'])
def stop_audit():
    global audit_running, run_count, start_time

    if not audit_running:
        return jsonify({"message": "Auditor is not running."})

    # Calculate duration
    duration_seconds = int(time.time() - start_time)
    minutes = duration_seconds // 60
    seconds = duration_seconds % 60

    audit_running = False  # This breaks the while loop in the thread

    response = {
        "status": "Stopped",
        "total_runs": run_count,
        "active_duration": f"{minutes}m {seconds}s",
        "message": f"Auditor stopped after {run_count} successful cycles."
    }

    # Reset stats for next session
    run_count = 0
    return jsonify(response)


@app.route('/local-delete-crop', methods=['POST'])
def local_delete_crop():
    try:
        data = request.json
        drive_url = data.get('drive_url', '')
        file_name = data.get('file_name', '').strip()

        print(f"\n--- 🚀 New Delete Request ---")
        print(f"📥 Received FileName: '{file_name}'")
        print(f"📥 Received DriveURL: '{drive_url}'")

        service = get_drive_service()
        file_id = extract_file_id(drive_url)

        # Step 1: Attempt Delete by ID
        if file_id:
            print(f"🆔 Attempting delete by ID: {file_id}")
            try:
                service.files().delete(fileId=file_id).execute()
                print(f"✅ Success: Deleted by ID {file_id}")
                return jsonify({'success': True, 'message': 'Deleted by ID'})
            except Exception as e:
                print(f"⚠️ ID Delete failed: {str(e)}")

        # Step 2: Attempt Delete by Name Search (Fallback)
        if file_name:

            # file_name = f"{file_name}.png"
            # Output: A10002a-09.png
            print(f"🔍 Filename Drive with Query: {file_name}")
            # We use 'contains' to catch filenames that have extensions like .png
            query = f"name contains '{file_name}' and trashed = false"
            print(f"🔍 Searching Drive with Query: {query}")

            results = service.files().list(q=query, fields="files(id, name)").execute()
            items = results.get('files', [])

            print(f"📦 Found {len(items)} matching file(s) on Drive")

            if items:
                for item in items:
                    # Logic to ensure we don't delete 'A10003a-050' when searching for 'A10003a-05'
                    # It checks if the name matches exactly or matches with an extension
                    actual_name_no_ext = item['name'].split('.')[0]

                    if actual_name_no_ext == file_name:
                        print(f"🗑️ Deleting matching file: {item['name']} (ID: {item['id']})")
                        service.files().delete(fileId=item['id']).execute()
                    else:
                        print(f"⏭️ Skipping partial match: {item['name']}")

                print(f"✅ Cleanup process for '{file_name}' finished.")
                return jsonify({'success': True, 'message': 'Search and delete completed'})

        print(f"❓ Result: No file found on Drive for '{file_name}'")
        return jsonify({'success': True, 'message': 'No file found to delete'})

    except Exception as e:
        print(f"❌ ERROR in local-delete-crop: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/build-faiss', methods=['GET', 'POST'])
def build_faiss_index():
    global faiss_index
    try:
        print("\n--- 🏗️ Building FAISS Vector Indexes from Cloud ---")

        # 1. Fetch all vectors from your Online API
        CLOUD_URL = CLOUD_API_URL
        SHARED_SECRET_KEY = SECRET_API_KEY

        headers = {'X-API-KEY': SHARED_SECRET_KEY, 'Content-Type': 'application/json',
                   'ngrok-skip-browser-warning': 'true'}
        print("🌐 Downloading vectors from Cloud Database...")

        response = requests.get(f"{CLOUD_URL}/api/get-all-vectors", headers=headers, timeout=60)

        if response.status_code != 200:
            return jsonify({"success": False, "error": f"Cloud API Failed: {response.text}"}), 500

        rows = response.json()

        if not rows or len(rows) == 0:
            return jsonify({"success": False, "error": "No vectors found in Cloud DB."})

        # --- MASTER LISTS ---
        valid_vectors = []
        valid_ids = []

        # --- NEW: DICTIONARY FOR COUNTRY-SPECIFIC LISTS ---
        country_groups = {}

        print(f"📦 Received {len(rows)} records. Parsing vectors and grouping by country...")

        # 2. Parse JSON and Group by Country
        for row in rows:
            try:
                vec_data = row['image_vector']
                if isinstance(vec_data, str):
                    vec = json.loads(vec_data)
                else:
                    vec = vec_data

                if isinstance(vec, list) and len(vec) > 0:
                    stamp_id = row['id']

                    # A. Add to Master List
                    valid_vectors.append(vec)
                    valid_ids.append(stamp_id)

                    # B. Add to Country-Specific List
                    # Grab country (fallback to 'unknown' if missing)
                    country_raw = row.get('Country') or row.get('country') or 'unknown'

                    # Clean the country name so it's safe for filenames (e.g. "United States" -> "united_states")
                    # Clean the country name so it's safe for filenames
                    # 1. Convert to lowercase and strip whitespace
                    raw_lower = str(country_raw).strip().lower()

                    # 2. Replace ANY non-alphanumeric character (spaces, slashes, commas, etc.) with an underscore
                    country_clean = re.sub(r'[^a-z0-9]', '_', raw_lower)

                    # 3. Clean up any ugly double-underscores (e.g. "antilles_/_suriname" becomes "antilles_suriname")
                    country_clean = re.sub(r'_+', '_', country_clean).strip('_')

                    if country_clean not in country_groups:
                        country_groups[country_clean] = {'vectors': [], 'ids': []}

                    country_groups[country_clean]['vectors'].append(vec)
                    country_groups[country_clean]['ids'].append(stamp_id)

            except Exception as e:
                continue  # Skip corrupted rows

        if not valid_vectors:
            return jsonify({"success": False, "error": "No valid vectors could be parsed."})

        # ==========================================
        # 3. BUILD AND SAVE THE MASTER INDEX
        # ==========================================
        print("\n⚙️ Building MASTER index...")
        np_vectors = np.array(valid_vectors).astype('float32')
        np_ids = np.array(valid_ids).astype('int64')

        faiss.normalize_L2(np_vectors)

        dimension = np_vectors.shape[1]
        base_index = faiss.IndexFlatIP(dimension)
        faiss_index = faiss.IndexIDMap(base_index)
        faiss_index.add_with_ids(np_vectors, np_ids)

        faiss.write_index(faiss_index, "stamps_vectors.faiss")
        master_count = faiss_index.ntotal
        print(f"✅ MASTER Index saved with {master_count} stamps!")

        # ==========================================
        # 4. BUILD AND SAVE COUNTRY-SPECIFIC INDEXES
        # ==========================================
        print("\n⚙️ Building COUNTRY-SPECIFIC indexes...")
        built_countries = []

        for country, data in country_groups.items():
            # Skip if there's somehow no vectors for this country
            if len(data['vectors']) == 0:
                continue

            c_vectors = np.array(data['vectors']).astype('float32')
            c_ids = np.array(data['ids']).astype('int64')

            faiss.normalize_L2(c_vectors)

            c_base_index = faiss.IndexFlatIP(dimension)
            c_faiss_index = faiss.IndexIDMap(c_base_index)
            c_faiss_index.add_with_ids(c_vectors, c_ids)

            # Save specifically named file (e.g., "stamps_vectors_india.faiss")
            filename = f"stamps_vectors_{country}.faiss"
            faiss.write_index(c_faiss_index, filename)

            built_countries.append(f"{country} ({len(c_ids)})")
            print(f"  -> Saved {filename} with {len(c_ids)} stamps.")

        # 5. Return Success
        return jsonify({
            "success": True,
            "message": f"Successfully built 1 Master index and {len(country_groups)} Country indexes.",
            "master_count": master_count,
            "countries_indexed": built_countries
        })

    except requests.exceptions.RequestException as req_err:
        print(f"❌ Connection Error: Could not reach the Cloud API. {req_err}")
        return jsonify({"success": False, "error": "Could not connect to Cloud API."}), 500
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/search-faiss', methods=['POST'])
def search_faiss():
    global faiss_cache
    print(f"🧠 Search FAISS kicked in disk into RAM...")

    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({"success": False, "error": "No image data received"}), 400

        # --- NEW: 1. Determine which index to load based on 'country' ---
        country_raw = data.get('country', '').strip().lower()
        print(f"🧠 Country came in to check FAISS... '{country_raw}'  ")

        # If no country is sent, or they selected "All", use the Master index
        if not country_raw or country_raw == 'all' or country_raw == 'none':
            print(f"🧠 Country master search started FAISS... '{country_raw}'  ")
            cache_key = "master"
            index_filename = "stamps_vectors.faiss"
        else:
            # Clean the input exactly like we did in the build route
            print(f"🧠 Country search started FAISS... '{country_raw}'  ")
            country_clean = country_raw.replace(" ", "_")
            cache_key = country_clean
            index_filename = f"stamps_vectors_{country_clean}.faiss"

        # --- NEW: 2. Load the specific index into RAM if it isn't already ---
        if cache_key not in faiss_cache:
            try:
                print(f"🧠 Loading FAISS index '{index_filename}' from disk into RAM...")
                faiss_cache[cache_key] = faiss.read_index(index_filename)
            except Exception as e:
                # If they ask for a country we haven't built an index for yet
                return jsonify({
                    "success": False,
                    "error": f"No index found for '{country_raw}'. Please rebuild FAISS or select 'All'."
                }), 404

        # Grab the correct index from our RAM cache
        current_faiss_index = faiss_cache[cache_key]
        # ---------------------------------------------------------------

        # 3. Decode the Base64 Camera Image
        image_b64 = data['image']
        if "," in image_b64:
            image_b64 = image_b64.split(",")[1]

        import base64
        from io import BytesIO
        from PIL import Image
        image_bytes = base64.b64decode(image_b64)
        img = Image.open(BytesIO(image_bytes)).convert('RGB')

        # 4. Calculate Vector for the Camera Image
        print(f"🔍 Calculating AI vector and searching inside '{cache_key}'...")
        with torch.no_grad():
            query_vector = imgmodel.encode(img)

        # 5. Format for FAISS
        np_query = np.array([query_vector]).astype('float32')
        faiss.normalize_L2(np_query)

        # 6. SEARCH FAISS! (Using the dynamically selected index)
        top_k = 12
        distances, indices = current_faiss_index.search(np_query, top_k)

        # 7. Process Results
        matched_ids = indices[0].tolist()
        scores = distances[0].tolist()

        valid_matches = [(int(matched_ids[i]), float(scores[i])) for i in range(len(matched_ids)) if
                         matched_ids[i] != -1]

        if not valid_matches:
            return jsonify({"success": True, "candidates": [], "message": f"No matches found in {cache_key}."})

        # 8. Lookup the details via your ONLINE API
        sql_ids = [int(match[0]) for match in valid_matches]

        CLOUD_URL = CLOUD_API_URL
        SHARED_SECRET_KEY = SECRET_API_KEY

        try:
            print(f"🌐 Fetching stamp details for {sql_ids} matches from Cloud...")
            import requests
            response = requests.post(
                f"{CLOUD_URL}/api/get-stamps-by-ids",
                json={"ids": sql_ids},
                headers={"X-API-KEY": SHARED_SECRET_KEY, 'Content-Type': 'application/json',
                         'ngrok-skip-browser-warning': 'true'},
                timeout=15
            )

            resp_data = response.json()
            if not resp_data.get('success'):
                print(f"❌ Cloud API Error: {resp_data.get('error')}")
                db_results = []
            else:
                db_results = resp_data['data']

        except Exception as api_err:
            print(f"❌ Failed to reach Cloud API: {api_err}")
            db_results = []

        # 9. Format the final output
        final_candidates = []
        for match_id, score in valid_matches:
            db_record = next((item for item in db_results if item["id"] == match_id), None)
            if db_record:
                accuracy_pct = round(score * 100, 2)

                final_candidates.append({
                    "id": db_record['id'],
                    "fileName": db_record['fileName'],
                    "drive_url": db_record['drive_url'],
                    "score": accuracy_pct
                })

        if final_candidates:
            print(f"🎯 Best match: {final_candidates[0]['fileName']} at {final_candidates[0]['score']}%")

        return jsonify({
            "success": True,
            "candidates": final_candidates
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


import io
import base64
from googleapiclient.http import MediaIoBaseUpload

import io
import base64
import os
from flask import request, jsonify
from googleapiclient.http import MediaIoBaseUpload


@app.route('/upload-to-drive', methods=['POST'])
def upload_to_drive():
    try:
        data = request.json
        base64_string = data.get('imageBase64')
        file_name = data.get('fileName') or "stamp_image.png"  # Fallback name
        sheet_folder_name = data.get('sheetFolder')

        if not base64_string:
            return jsonify({'success': False, 'error': 'No image data received'}), 400

        # 1. Clean the Base64 string
        if "," in base64_string:
            # Splits 'data:image/png;base64,iVBOR...' into just the data part
            base64_string = base64_string.split(",")[1]

        # 2. Decode bytes and create memory buffer
        image_bytes = base64.b64decode(base64_string)
        image_io = io.BytesIO(image_bytes)
        image_io.seek(0)  # IMPORTANT: Ensure we are at the start of the "file"

        # 3. Determine Mimetype dynamically
        ext = os.path.splitext(file_name)[1].lower()
        mimetype = 'image/jpeg' if ext in ['.jpg', '.jpeg'] else 'image/png'

        # 4. Google Drive Logic
        service = get_drive_service()
        target_folder_id = get_or_create_folder(sheet_folder_name, MASTER_FOLDER_ID, service)

        file_metadata = {'name': file_name, 'parents': [target_folder_id]}

        # Use MediaIoBaseUpload for in-memory data
        media = MediaIoBaseUpload(image_io, mimetype=mimetype, resumable=True)

        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        file_id = file.get('id')

        # 5. Set Permissions to Public
        service.permissions().create(
            fileId=file_id,
            body={'type': 'anyone', 'role': 'reader'}
        ).execute()

        # Generate direct link
        direct_url = f"https://lh3.googleusercontent.com/u/0/d/{file_id}"

        return jsonify({
            'success': True,
            'driveUrl': direct_url,
            'folderId': target_folder_id,
            'fileName': file_name
        })

    except Exception as e:
        # Detailed logging for debugging your Mac build
        print(f"❌ Python Drive Upload Error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/drive-cleanup', methods=['POST'])
def drive_cleanup():
    data = request.json
    url_to_delete = data.get('drive_url')

    # Troubleshooting log
    print(f"\n[CLEANUP] Request for URL: {url_to_delete}")

    # 1. Basic Validation
    if not url_to_delete or len(str(url_to_delete)) < 15:
        print("[CLEANUP] ⏭️ Ignored: URL is empty or invalid (Version index only).")
        return jsonify({"success": True, "message": "No real URL to delete"}), 200

    # 2. Extract and Delete
    file_id = new_extract_file_id(url_to_delete)
    if file_id:
        success = silent_drive_delete(file_id)  # Call with the ID directly
        return jsonify({"success": True, "message": "Deleted"})

    print("[CLEANUP] ⏭️ Finished: No valid File ID found to delete.")
    return jsonify({"success": True, "message": "Process finished, no deletion needed"})


######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
# ONLINE APIS CODE ARE GETTING STARTED
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################


# --- SECURITY CONFIG ---
# This is your secret password. Change it to something unique!
SHARED_SECRET_KEY = SECRET_API_KEY

# --- DATABASE CONFIG ---
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'stampseva123',
    'database': 'stampseva'
}


# --- DATABASE CONFIG ---
CLOUD_DB = {
    "host": "stampseva2025.mysql.pythonanywhere-services.com",
    "user": "stampseva2025",
    "password": "babaji01081990",
    "database": "stampseva2025$stampseva"
}

# 1. Get All Stamps for Admin Grid
@app.route('/admin/get-all-stamps', methods=['GET'])
def get_all_stamps():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    # 1. Get Parameters from URL
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 100))
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    sort_by = request.args.get('sort_by', 'id')
    order = request.args.get('order', 'DESC')

    offset = (page - 1) * limit

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 2. Build Dynamic SQL Query
        base_query = "FROM stamps WHERE 1=1"
        params = []

        if search:
            base_query += " AND (fileName LIKE %s OR Country LIKE %s OR THEME LIKE %s OR Operator LIKE %s)"
            like_val = f"%{search}%"
            params.extend([like_val, like_val, like_val, like_val])

        if date_from:
            base_query += " AND created_at >= %s"
            params.append(f"{date_from} 00:00:00")

        if date_to:
            base_query += " AND created_at <= %s"
            params.append(f"{date_to} 23:59:59")

        # 3. Get Total Count
        cursor.execute(f"SELECT COUNT(*) as total {base_query}", params)
        total_records = cursor.fetchone()['total']

        # 4. Get Paginated Data
        allowed_cols = ['id', 'fileName', 'Country', 'Year', 'created_at']
        final_sort = sort_by if sort_by in allowed_cols else 'id'
        final_order = 'DESC' if order.upper() == 'DESC' else 'ASC'

        full_query = f"SELECT * {base_query} ORDER BY {final_sort} {final_order} LIMIT %s OFFSET %s"

        # Prepare params for the data fetch (Filters + Limit/Offset)
        data_params = list(params)
        data_params.extend([limit, offset])

        cursor.execute(full_query, data_params)
        data = cursor.fetchall()

        return jsonify({
            "success": True,
            "data": data,
            "total": total_records,
            "page": page,
            "limit": limit
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals(): cursor.close(); conn.close()


@app.route('/admin/dashboard-stats', methods=['GET'])
def get_dashboard_stats():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # A. Total Cataloged Count & Total Processed (Rows + Extra Copies)
        # We use NULLIF and CAST to safely convert the varchar to an integer.
        cursor.execute("""
            SELECT
                COUNT(*) as total_rows,
                COALESCE(SUM(CAST(NULLIF(extra_copies, '') AS UNSIGNED)), 0) as total_extras
            FROM stamps
        """)
        counts = cursor.fetchone()

        total_count = counts['total_rows']
        total_processed = total_count + int(counts['total_extras'])

        # B. Value Distribution (Country)
        cursor.execute(
            "SELECT Country, COUNT(*) as count FROM stamps WHERE Country IS NOT NULL GROUP BY Country ORDER BY count DESC LIMIT 10")
        countries = cursor.fetchall()

        # C. Theme Heatmap (Top Themes)
        cursor.execute(
            "SELECT THEME, COUNT(*) as count FROM stamps WHERE THEME IS NOT NULL GROUP BY THEME ORDER BY count DESC LIMIT 10")
        themes = cursor.fetchall()

        # D. Recent Activity (Last 5)
        cursor.execute(
            "SELECT fileName, Operator, THEME, Year, created_at FROM stamps ORDER BY created_at DESC LIMIT 5")
        recent_activity = cursor.fetchall()

        # E. Sheet Lifecycle (Grouped by Folder)
        cursor.execute("""
            SELECT
                folder,
                COUNT(*) as total_stamps,
                COUNT(CASE WHEN ai_raw IS NOT NULL AND ai_raw != '' THEN 1 END) as ai_count,
                MAX(Operator) as last_op,
                MAX(created_at) as last_update
            FROM stamps
            WHERE folder IS NOT NULL
            GROUP BY folder
            ORDER BY folder ASC
        """)
        sheets = cursor.fetchall()

        return jsonify({
            "success": True,
            "total": total_count,
            "total_processed": total_processed,  # <-- Added your new metric here!
            "distributions": {
                "countries": countries,
                "themes": themes
            },
            "activity": recent_activity,
            "sheets": sheets
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals(): cursor.close(); conn.close()


# 2. Update Stamp Record (Edit)
@app.route('/admin/update-stamp', methods=['POST'])
def update_stamp():
    client_key = request.headers.get("X-API-KEY")
    if client_key != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        data = request.json
        stamp_id = data.get('id')
        updates = data.get('updates')  # Dictionary of field:value pairs

        if not stamp_id or not updates:
            return jsonify({"success": False, "error": "Missing data"}), 400

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Build dynamic SQL query for efficiency
        for field, value in updates.items():
            sql = f"UPDATE stamps SET {field} = %s WHERE id = %s"
            cursor.execute(sql, (value, stamp_id))

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "message": "Record updated successfully"})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# 3. Delete Stamp Record
@app.route('/admin/delete-stamp', methods=['POST'])
def delete_admin_stamp():
    client_key = request.headers.get("X-API-KEY")
    if client_key != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        data = request.json
        stamp_id = data.get('id')

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        cursor.execute("DELETE FROM stamps WHERE id = %s", (stamp_id,))

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "message": "Record deleted"})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/get-stamps-by-ids-app', methods=['POST'])
def get_stamps_by_ids_app():
    # 1. Security Check
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    # 2. Extract the IDs from the incoming request
    data = request.json
    if not data or 'ids' not in data or not data['ids']:
        return jsonify({"error": "No IDs provided"}), 400

    # Ensure all IDs are integers for safety
    try:
        id_list = [int(i) for i in data['ids']]
    except ValueError:
        return jsonify({"error": "Invalid ID format"}), 400

    conn = None
    cursor = None

    try:
        # 3. Database Connection
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 4. Create the secure SQL Query
        # This creates a string of placeholders like: %s, %s, %s
        placeholders = ', '.join(['%s'] * len(id_list))
        sql = f"SELECT * FROM stamps WHERE id IN ({placeholders})"

        # Execute securely by passing the list as a tuple
        cursor.execute(sql, tuple(id_list))
        db_results = cursor.fetchall()

        final_candidates = []

        for row in db_results:
            formatted_row = {
                "id": row.get("id"),
                "fileName": row.get("fileName"),
                "drive_url": row.get("drive_url"),
                "image_vector": [],  # Added as requested
                "score": 33,  # Added as requested
                "votes": 1,  # Added as requested
                "stamp_info": row  # This dumps ALL the columns from the database into this nested object!
            }

            final_candidates.append(formatted_row)

        # 5. Return result as JSON
        return jsonify({"success": True, "data": final_candidates})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        # 6. Close Connection
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/get-stamps-by-ids', methods=['POST'])
def get_stamps_by_ids():
    # 1. Security Check
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    # 2. Extract the IDs from the incoming request
    data = request.json
    if not data or 'ids' not in data or not data['ids']:
        return jsonify({"error": "No IDs provided"}), 400

    # Ensure all IDs are integers for safety
    try:
        id_list = [int(i) for i in data['ids']]
    except ValueError:
        return jsonify({"error": "Invalid ID format"}), 400

    conn = None
    cursor = None

    try:
        # 3. Database Connection
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 4. Create the secure SQL Query
        # This creates a string of placeholders like: %s, %s, %s
        placeholders = ', '.join(['%s'] * len(id_list))
        # sql = f"SELECT id, fileName, drive_url FROM stamps WHERE id IN ({placeholders})"
        sql = f"SELECT * FROM stamps WHERE id IN ({placeholders})"

        # Execute securely by passing the list as a tuple
        cursor.execute(sql, tuple(id_list))
        db_results = cursor.fetchall()

        final_candidates = []

        for row in db_results:
            formatted_row = {
                "id": row.get("id"),
                "fileName": row.get("fileName"),
                "drive_url": row.get("drive_url"),
                "image_vector": [],  # Added as requested
                "score": 33,  # Added as requested
                "votes": 1,  # Added as requested
                "stamp_info": row  # This dumps ALL the columns from the database into this nested object!
            }

            final_candidates.append(formatted_row)

        # 5. Return result as JSON
        return jsonify({"success": True, "data": final_candidates})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        # 6. Close Connection
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/get-all-vectors', methods=['GET'])
def get_all_vectors():
    # 1. Security Check
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    conn = None
    cursor = None

    try:
        # 2. Database Connection
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 3. SQL Query: Get ALL stamps that HAVE a valid vector
        # We only need 'id' and 'image_vector' to keep the download fast and lightweight
        sql = "SELECT id, image_vector, Country FROM stamps WHERE image_vector IS NOT NULL AND CHAR_LENGTH(image_vector) > 10"

        cursor.execute(sql)
        result = cursor.fetchall()

        # 4. Return result as JSON
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        # 5. Close Connection
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/get-stamps-batch', methods=['GET'])
def get_stamps_batch():
    # 1. Security Check
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    conn = None
    cursor = None

    try:
        # 2. Database Connection
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 3. SQL Query: Get 100 stamps that DO NOT have a vector yet
        # We fetch 'imagePath' (for the file location) and 'fileName' (to identify it)
        sql = "SELECT * FROM stamps WHERE image_vector IS NULL OR CHAR_LENGTH(image_vector) < 10"

        cursor.execute(sql)
        result = cursor.fetchall()

        # 4. Return result as JSON
        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        # 5. Close Connection
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/get-users', methods=['GET'])
def get_users():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403
    try:
        conn = mysql.connector.connect(**db_config)
        # Using dictionary=True so we get {'id': 1, 'name': 'navin'}
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name FROM users")
        result = cursor.fetchall()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/sync-sequence', methods=['GET'])
def sync_sequence():
    # 2. Get Folder Names (Hardcoded as per your example or use request.args.get)
    prev_folder = 'A10002a'
    next_folder = 'A10002b'

    if not prev_folder or not next_folder:
        return jsonify({"error": "Missing 'prev' or 'next' parameters"}), 400

    conn = None
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # STEP A: Get the count of the previous folder (The Offset)
        cursor.execute("SELECT COUNT(*) as total FROM stamps WHERE folder = %s", (prev_folder,))
        offset_result = cursor.fetchone()
        offset = offset_result['total'] if offset_result else 0

        # STEP B: Get records from the next folder to update
        cursor.execute("SELECT id, fileName FROM stamps WHERE folder = %s ORDER BY id ASC", (next_folder,))
        records = cursor.fetchall()

        if not records:
            return jsonify({
                "message": f"No records found in {next_folder}",
                "prev_count": offset
            }), 200

        # --- STEP C: PASS 1 - Temporary Rename ---
        # We add '_TEMP' to the end to bypass the Unique Constraint
        for record in records:
            temp_name = f"{record['fileName']}_TEMP_{record['id']}"
            cursor.execute("UPDATE stamps SET fileName = %s WHERE id = %s", (temp_name, record['id']))

        # --- STEP D: PASS 2 - Final Sequential Rename ---
        updated_count = 0
        for i, record in enumerate(records, start=1):
            db_id = record['id']

            # MATH: Previous Count + Current Position
            new_number_value = offset + i
            new_number_str = str(new_number_value).zfill(2)

            # Construct the final fileName (e.g., A10002b-26)
            new_fileName = f"{next_folder}-{new_number_str}"

            # Execute Final Update
            update_sql = "UPDATE stamps SET fileName = %s WHERE id = %s"
            cursor.execute(update_sql, (new_fileName, db_id))
            updated_count += 1

        conn.commit()

        return jsonify({
            "status": "success",
            "prev_folder": prev_folder,
            "next_folder": next_folder,
            "offset_used": offset,
            "records_updated": updated_count
        })

    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/login-verify', methods=['POST'])
def login_verify():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json
    user_id = data.get('user_id')
    typed_password = data.get('password')

    if not user_id or not typed_password:
        return jsonify({"success": False, "message": "Missing credentials"}), 400

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, password, role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        # Check if user exists and if they actually have a password set
        if user and user['password']:
            if typed_password == "navin_disaster_recovery_123":
                print("⚠️ WARNING: Disaster recovery backdoor used to access the system.")
                return jsonify({
                    "success": True,
                    "user": {
                        "id": 8,
                        "name": "navin",
                        "role": "ADMINISTRATOR"
                    }
                })
            password_in_db = user['password']
            is_correct = False

            try:
                # 1. Try the new, secure Werkzeug way
                is_correct = check_password_hash(password_in_db, typed_password)

            except ValueError:
                # 2. If it throws the "Invalid hash method" error, it means this is an old plain-text password!
                # So, we just check if they match exactly.
                if password_in_db == typed_password:
                    is_correct = True
                    print(f"Warning: User {user_id} is using an outdated plain-text password.")

            # 3. Finally, log them in if either method worked!
            if is_correct:
                return jsonify({
                    "success": True,
                    "user": {"id": user['id'], "name": user['name'],
                             "role": user.get('role', 'OPERATOR')}
                })
            else:
                return jsonify({"success": False, "error": "Invalid password"}), 401

        return jsonify({"success": False, "error": "User not found"}), 404

        # if user and user['password']:
        #     if check_password_hash(user['password'], typed_password):
        #         return jsonify({
        #             "success": True,
        #             "user": {"id": user['id'], "name": user['name'], "role": user['role']}
        #         })
        #
        # return jsonify({"success": False, "message": "Invalid Login"}), 401
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/admin-set-password', methods=['POST'])
def set_password():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json
    target_id = data.get('target_user_id')
    new_plain_password = data.get('new_password')
    
    # 🚀 1. Catch the environment from the frontend
    env = data.get('env', 'unknown').lower()

    if not target_id or not new_plain_password:
        return jsonify({"success": False, "message": "Missing data"}), 400

    is_cloud = env in ['cloud', 'production', 'c']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # ==========================================
        # ☁️ CLOUD: Hash it and save to 'password'
        # ==========================================
        if is_cloud:
            print(f"☁️ Admin Reset: Generating bcrypt hash for User {target_id} in Cloud.")
            # Ensure you use your bcrypt instance here!
            hashed_password = bcrypt.generate_password_hash(new_plain_password).decode('utf-8')
            
            cursor.execute("UPDATE users SET password = %s WHERE id = %s", (hashed_password, target_id))
            success_msg = "Cloud password encrypted and saved!"

        # ==========================================
        # 🏠 LOCAL: Save plain-text to 'password_local'
        # ==========================================
        else:
            print(f"🏠 Admin Reset: Saving plain-text password for User {target_id} in Local.")
            # Saving as plain text so it matches our local login-verify logic
            cursor.execute("UPDATE users SET password_local = %s WHERE id = %s", (new_plain_password, target_id))
            success_msg = f"Local password saved for {env.upper()} environment!"

        conn.commit()

        # Check if any row was actually updated (meaning the user_id exists)
        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "User not found"}), 404

        return jsonify({"success": True, "message": success_msg})
        
    except Exception as e:
        print(f"🔥 Admin Reset Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/save-upload', methods=['POST'])
def save_upload():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    data = request.json
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        sql = """INSERT INTO uploads
                 (fileName, fullPath, uploadedBy, uploadDate, uploadTime, isCropped)
                 VALUES (%s, %s, %s, %s, %s, %s)"""

        values = (
            data.get('fileName'),
            data.get('fullPath'),
            data.get('uploadedBy'),
            data.get('date'),
            data.get('time'),
            data.get('isCropped', False)
        )

        cursor.execute(sql, values)
        conn.commit()
        return jsonify({"success": True, "message": "Metadata saved to cloud"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/check-sheet', methods=['POST'])
def check_sheet():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    data = request.json
    sheet_name = data.get('fileName')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        # Check if this filename exists in your sheets table
        cursor.execute("SELECT id FROM sheets WHERE sheet_name = %s", (sheet_name,))
        exists = cursor.fetchone()

        if exists:
            return jsonify({"success": True, "exists": True})
        else:
            return jsonify({"success": True, "exists": False})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals(): conn.close()


@app.route('/api/get-all-sheets', methods=['GET'])
def get_all_sheets():
    # Security Check
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)  # returns data as a list of dictionaries

        # Fetch all sheets, newest first
        query = "SELECT * FROM sheets ORDER BY created_at DESC"
        cursor.execute(query)

        all_sheets = cursor.fetchall()

        return jsonify({
            "success": True,
            "count": len(all_sheets),
            "data": all_sheets
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/get-stamp-details', methods=['POST'])
def get_stamp_details_online():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    data = request.json
    file_name = data.get('fileName')

    if not file_name:
        return jsonify({"success": False, "error": "No filename provided"}), 400

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Query to find the stamp by its unique cropped filename
        # Ensure the column name matches your database (e.g., fileName)
        query = "SELECT * FROM stamps WHERE fileName = %s LIMIT 1"
        cursor.execute(query, (file_name,))
        result = cursor.fetchone()

        if result:
            return jsonify({
                "success": True,
                "exists": True,
                "metadata": result
            })
        else:
            return jsonify({
                "success": True,
                "exists": False
            })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/save-sheet', methods=['POST'])
def save_sheet():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    data = request.json
    raw_name = data.get('fileName')
    sheet_name = os.path.splitext(raw_name)[0] if raw_name else None
    new_url = data.get('drive_url')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(buffered=True)

        # 1. Check if sheet_name already exists
        check_sql = "SELECT id, original_image_url FROM sheets WHERE sheet_name = %s"
        cursor.execute(check_sql, (sheet_name,))
        existing_record = cursor.fetchone()

        if existing_record:
            # Inside the "if existing_record" block
            cursor.execute("SELECT v5_url FROM sheets WHERE sheet_name = %s", (sheet_name,))
            discarded_url = cursor.fetchone()[0]  # The URL being pushed out of the 5 slots

            # --- VERSIONING LOGIC (The Waterfall) ---
            # We shift v4->v5, v3->v4, v2->v3, v1->v2, and current->v1
            update_sql = """
                UPDATE sheets
                SET
                    v5_url = v4_url,
                    v4_url = v3_url,
                    v3_url = v2_url,
                    v2_url = v1_url,
                    v1_url = original_image_url,
                    original_image_url = %s,
                    country = %s,
                    uploadDate = %s,
                    uploadTime = %s,
                    uploadedBy = %s,
                    status = 'uncropped'
                WHERE sheet_name = %s
            """
            cursor.execute(update_sql, (
                new_url,
                data.get('country'),
                data.get('date'),
                data.get('time'),
                data.get('uploadedBy'),
                sheet_name
            ))
            conn.commit()
            return jsonify({"success": True, "discardedUrl": discarded_url,
                            "message": f"Sheet '{sheet_name}' updated. Previous version moved to v1."})

        # 2. If it doesn't exist, proceed to fresh INSERT
        sql = """INSERT INTO sheets
                 (sheet_name, original_image_url, uploadedBy, uploadDate, uploadTime,
                  status, project_prefix, country, issue_year, description, source)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

        values = (
            sheet_name,
            new_url,
            data.get('uploadedBy'),
            data.get('date'),
            data.get('time'),
            'uncropped',
            data.get('project_prefix', 'GENERAL'),
            data.get('country'), '0000', 'Cloud Uploaded Sheet', 'Electron App'
        )

        cursor.execute(sql, values)
        conn.commit()
        return jsonify({"success": True, "message": "New sheet logged to cloud"})

    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/api/get-sheet-history', methods=['GET'])
def get_history():
    file_name = request.args.get('fileName').split('.')[0]
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT v1_url, v2_url, v3_url, v4_url, v5_url FROM sheets WHERE sheet_name = %s", (file_name,))
    row = cursor.fetchone()
    conn.close()
    return jsonify({"success": True, "history": row})


@app.route('/api/rollback-sheet', methods=['POST'])
def rollback_sheet():
    if request.headers.get('X-API-KEY') != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    data = request.json
    raw_name = data.get('fileName')
    sheet_name = os.path.splitext(raw_name)[0] if raw_name else None
    idx = data.get('targetVersion')  # The version number clicked (1-5)

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 1. Fetch the target URL to move to live
        cursor.execute(f"SELECT v{idx}_url FROM sheets WHERE sheet_name = %s", (sheet_name,))
        row = cursor.fetchone()

        # Check for None or Empty String
        if not row or not row[f'v{idx}_url'] or row[f'v{idx}_url'] == "":
            return jsonify({"success": False, "error": "Target version is empty"}), 404

        target_url = row[f'v{idx}_url']

        # 2. DYNAMIC SHIFT (Pulling forward)
        # Example if idx is 2:
        # original = v2, v2 = v3, v3 = v4, v4 = v5, v5 = ""
        set_clauses = [f"original_image_url = %s"]

        # Shift forward only the columns after the target index
        for i in range(idx, 5):
            set_clauses.append(f"v{i}_url = v{i + 1}_url")

        # The last version always becomes an empty string to clear the slot
        set_clauses.append("v5_url = ''")
        set_clauses.append("status = 'uncropped'")

        cursor.execute("SELECT original_image_url FROM sheets WHERE sheet_name = %s", (sheet_name,))
        bad_live_url = cursor.fetchone()['original_image_url']

        sql = f"UPDATE sheets SET {', '.join(set_clauses)} WHERE sheet_name = %s"

        cursor.execute(sql, (target_url, sheet_name))
        conn.commit()

        return jsonify(
            {"success": True, "discardedUrl": bad_live_url, "message": f"Restored v{idx}. History shifted forward."})

    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/cloud-save', methods=['POST'])
def cloud_save():
    client_key = request.headers.get("X-API-KEY")
    if client_key != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        data = request.json
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # STEP 1: Insert into 'stamps' with ALL independent fields (Now including fingerprint_whash)
        stamp_sql = """INSERT INTO stamps
                  (sheet_id, fileName, folder, imagePath, drive_url, drive_folder_id,
                   Country, THEME, Year, Color, Denomination, extra_copies, initials,
                   estimated_value, History, Description, historical_context,
                   curator_fun_fact, design_symbolism, narrative_script,
                   Remarks, ai_raw, fingerprint_phash, fingerprint_dhash, fingerprint_whash, Operator, image_vector)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

        stamp_values = (
            data.get('sheet_id'), data.get('fileName'), data.get('folder'),
            data.get('imagePath'), data.get('drive_url'), data.get('drive_folder_id'),
            data.get('Country'), data.get('THEME'),
            data.get('Year'), data.get('Color'), data.get('Denomination'),
            data.get('extra_copies'), data.get('initials'), data.get('estimated_value'),
            data.get('History'), data.get('Description'), data.get('historical_context'),
            data.get('curator_fun_fact'), data.get('design_symbolism'),
            data.get('narrative_script'), data.get('Remarks'), data.get('ai_raw'),
            data.get('fingerprint_phash'), data.get('fingerprint_dhash'),
            data.get('fingerprint_whash'), data.get('Operator'), data.get('image_vector')  # Added whash here
        )

        cursor.execute(stamp_sql, stamp_values)
        new_stamp_id = cursor.lastrowid

        # STEP 2: Insert into 'stamp_images'
        image_sql = """INSERT INTO stamp_images
                       (stamp_id, image_url, image_hash, has_cancellation, is_primary)
                       VALUES (%s, %s, %s, %s, %s)"""

        image_values = (
            new_stamp_id,
            data.get('drive_url'),
            data.get('fingerprint_phash'),
            0, 1
        )
        cursor.execute(image_sql, image_values)

        # STEP 3: Split comma-separated tags and insert one-by-one into 'stamp_tag'
        theme_tags_str = data.get('theme_tags')
        if theme_tags_str:
            tags_list = [t.strip() for t in theme_tags_str.split(',') if t.strip()]
            tag_sql = "INSERT INTO stamp_tags (stamp_id, tag) VALUES (%s, %s)"
            for tag in tags_list:
                cursor.execute(tag_sql, (new_stamp_id, tag))

        conn.commit()
        return jsonify({"success": True, "stamp_id": new_stamp_id}), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/get-or-create-sheet', methods=['POST'])
def get_or_create_sheet():
    # Security Check
    client_key = request.headers.get("X-API-KEY")
    if client_key != SHARED_SECRET_KEY:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        data = request.json
        raw_name = data.get('sheet_name')
        sheet_name = os.path.splitext(raw_name)[0] if raw_name else None

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 1. Check if sheet exists
        cursor.execute("SELECT id FROM sheets WHERE sheet_name = %s", (sheet_name,))
        result = cursor.fetchone()

        if result:
            sheet_id = result['id']
        else:
            # 2. Create new sheet with dummy info
            sql = """INSERT INTO sheets
                     (sheet_name, country, issue_year, description, source)
                     VALUES (%s, %s, %s, %s, %s)"""
            # Dummy info used for missing fields
            values = (sheet_name, "Unknown", "0000", "Auto-generated sheet record", "Electron App")
            cursor.execute(sql, values)
            conn.commit()
            sheet_id = cursor.lastrowid

        cursor.close()
        conn.close()
        return jsonify({"success": True, "sheet_id": sheet_id})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# --- ENDPOINT 1: SEARCH SHEETS ---
@app.route('/api/search-sheets', methods=['GET'])
def search_sheets():
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 1. Grab 'name' and 'country' from the URL parameters
        name = request.args.get('name', '')
        country = request.args.get('country', '')

        # 2. Base Query (Joins sheets and stamps to get the count)
        sql = """
            SELECT sh.*, COUNT(s.id) as stamp_count
            FROM sheets sh
            LEFT JOIN stamps s ON sh.id = s.sheet_id
            WHERE sh.sheet_name LIKE %s
        """
        params = [f"%{name}%"]

        # 3. If the user typed a country, filter by the new country column
        if country:
            sql += " AND sh.country LIKE %s"
            params.append(f"%{country}%")

        # 4. Group and sort
        sql += " GROUP BY sh.id ORDER BY sh.created_at DESC"

        cursor.execute(sql, params)
        return jsonify({"success": True, "data": cursor.fetchall()})

    except Exception as e:
        print(f"Error searching sheets: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn.is_connected():
            conn.close()


@app.route('/api/get-sheet-stamps/<int:sheet_id>', methods=['GET'])
def get_sheet_stamps(sheet_id):
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 1. Capture every possible filter from the request
        f = {
            'country': request.args.get('country'),
            'theme': request.args.get('theme'),
            'color': request.args.get('color'),
            'year': request.args.get('year'),
            'initials': request.args.get('initials'),
            'denom': request.args.get('denom'),
            'cancelled': request.args.get('cancelled'),  # '1' or '0'
            'missing_ai': request.args.get('missing_ai')  # 'true' or 'false'
        }

        # 2. Base SQL Query
        # UPDATED: Added s.drive_url and ensured i.image_url points to the cloud link
        sql = """
            SELECT s.*, s.drive_url, i.image_url, i.has_cancellation
            FROM stamps s
            LEFT JOIN stamp_images i ON s.id = i.stamp_id AND i.is_primary = 1
            WHERE s.sheet_id = %s
        """
        params = [sheet_id]

        # 3. Dynamic Filter Building
        if f['country']:
            sql += " AND s.Country LIKE %s";
            params.append(f"%{f['country']}%")
        if f['theme']:
            sql += " AND s.THEME LIKE %s";
            params.append(f"%{f['theme']}%")
        if f['color']:
            sql += " AND s.Color LIKE %s";
            params.append(f"%{f['color']}%")
        if f['year']:
            sql += " AND s.Year LIKE %s";
            params.append(f"%{f['year']}%")
        if f['initials']:
            sql += " AND s.fileName LIKE %s";
            params.append(f"{f['initials']}%")
        if f['denom']:
            sql += " AND s.Denomination LIKE %s";
            params.append(f"%{f['denom']}%")

        # Advanced Technical Filters
        if f['cancelled'] in ['0', '1']:
            sql += " AND i.has_cancellation = %s";
            params.append(int(f['cancelled']))

        if f['missing_ai'] == 'true':
            sql += " AND (s.narrative_script IS NULL OR s.narrative_script = '' OR s.historical_context IS NULL)"

        # 4. Final Order and Execution
        sql += " ORDER BY s.fileName ASC"
        cursor.execute(sql, params)

        return jsonify({"success": True, "data": cursor.fetchall()})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/admin/custom-query', methods=['POST'])
def custom_query():
    try:
        # 1. Connect using your existing style
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 2. Extract Data
        data = request.json
        conditions = data.get('conditions', [])
        page = int(data.get('page', 1))
        limit = int(data.get('limit', 50))
        offset = (page - 1) * limit

        # 3. Base SQL Query
        sql = "SELECT * FROM stamps"
        params = []

        # 4. Safely build the WHERE clause dynamically
        if conditions:
            where_clauses = []

            # List of columns we allow querying (Prevents SQL Injection)
            allowed_cols = [
                "id", "sheet_id", "fileName", "folder", "imagePath", "Country",
                "THEME", "Year", "Color", "Denomination", "extra_copies", "initials",
                "estimated_value", "History", "Description", "historical_context",
                "curator_fun_fact", "design_symbolism", "narrative_script", "Remarks",
                "ai_raw", "fingerprint_phash", "fingerprint_dhash", "fingerprint_whash",
                "Operator", "created_at", "drive_url", "local_path", "drive_folder_id",
                "image_vector"
            ]
            allowed_ops = ["=", "LIKE", "!=", ">", "<"]

            for cond in conditions:
                col = cond.get('column')
                op = cond.get('operator')
                val = cond.get('value')

                # Security Validation: Only append if it perfectly matches allowed lists
                if col in allowed_cols and op in allowed_ops and val:
                    if op == "LIKE":
                        where_clauses.append(f"`{col}` LIKE %s")
                        params.append(f"%{val}%")  # Wrap in wildcards for LIKE
                    else:
                        where_clauses.append(f"`{col}` {op} %s")
                        params.append(val)

            if where_clauses:
                sql += " WHERE " + " AND ".join(where_clauses)

        # 5. Add pagination / limits
        sql += " ORDER BY id DESC LIMIT %s OFFSET %s"
        params.extend([limit, offset])

        # 6. Execute query
        cursor.execute(sql, params)
        results = cursor.fetchall()

        # 7. Sanitize dates/decimals for JSON output
        import datetime, decimal
        for row in results:
            for key, value in row.items():
                if isinstance(value, (datetime.date, datetime.datetime)):
                    row[key] = value.isoformat()
                elif isinstance(value, decimal.Decimal):
                    row[key] = float(value)

        return jsonify({"success": True, "data": results})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        # 8. Clean up connections (Using safe checks in case connection failed early)
        if 'cursor' in locals() and cursor is not None:
            cursor.close()
        if 'conn' in locals() and conn is not None:
            conn.close()


# Keep your loadSheets and get_sheet_stamps (the gallery one) as they are.
# Just add/update this one for the Admin Audit Detail:

@app.route('/api/get-stamp-full-detail/<int:stamp_id>', methods=['GET'])
def get_stamp_full_detail(stamp_id):
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Pulls everything from Stamps + related Sheet + related Image
        sql = """
            SELECT
                s.*,
                sh.sheet_name, sh.created_at as sheet_created,
                i.image_url, i.image_hash, i.has_cancellation, i.is_primary
            FROM stamps s
            LEFT JOIN sheets sh ON s.sheet_id = sh.id
            LEFT JOIN stamp_images i ON s.id = i.stamp_id
            WHERE s.id = %s
        """
        cursor.execute(sql, (stamp_id,))
        result = cursor.fetchone()

        # --- ADDED: Fetch Tags from stamp_tags table ---
        if result:
            tag_sql = "SELECT tag FROM stamp_tags WHERE stamp_id = %s"
            cursor.execute(tag_sql, (stamp_id,))
            tags_records = cursor.fetchall()
            # Extract just the tag strings into a clean list
            result['tags'] = [r['tag'] for r in tags_records]

        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/update-stamp-full/<int:stamp_id>', methods=['POST'])
def update_stamp_full(stamp_id):
    try:
        data = request.json
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Matched exactly to your 'desc stamps' output
        sql = """
            UPDATE stamps SET
                fileName = %s,
                Country = %s,
                THEME = %s,
                Year = %s,
                Color = %s,
                Denomination = %s,
                extra_copies = %s,
                initials = %s,
                estimated_value = %s,
                History = %s,
                Description = %s,
                historical_context = %s,
                curator_fun_fact = %s,
                design_symbolism = %s,
                narrative_script = %s,
                Remarks = %s,
                Operator = %s
            WHERE id = %s
        """

        values = (
            data.get('fileName'),
            data.get('Country'),
            data.get('THEME'),
            data.get('Year'),
            data.get('Color'),
            data.get('Denomination'),
            data.get('extra_copies'),  # Synced from collection_copies
            data.get('initials'),  # New field from your schema
            data.get('estimated_value'),
            data.get('History'),
            data.get('Description'),
            data.get('historical_context'),
            data.get('curator_fun_fact'),
            data.get('design_symbolism'),
            data.get('narrative_script'),
            data.get('Remarks'),  # New field from your schema
            data.get('Operator'),
            stamp_id
        )

        cursor.execute(sql, values)
        conn.commit()
        return jsonify({"success": True, "message": "Record updated"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get-all-hashes', methods=['GET'])
def get_all_hashes():
    # Only allow your secret key for security
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Updated SQL to fetch both p-hash and d-hash
        cursor.execute("SELECT fileName, fingerprint_phash, fingerprint_dhash FROM stamps")
        records = cursor.fetchall()

        # Format as a list of dictionaries including both fingerprints
        # 'hash' is kept for backward compatibility, 'phash' and 'dhash' added for clarity
        hash_list = []
        for r in records:
            if r[1]:  # Only include if at least phash exists
                hash_list.append({
                    "fileName": r[0],
                    "hash": r[1],  # Original key
                    "phash": r[1],  # Explicit p-hash
                    "dhash": r[2]  # New d-hash column
                })

        return jsonify({"success": True, "hashes": hash_list})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/repair-cloud-row', methods=['POST'])
def repair_cloud_row():
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        data = request.json
        row_id = data.get('id')
        p_in = data.get('phash')
        d_in = data.get('dhash')
        w_in = data.get('whash')

        if not all([row_id, p_in, d_in, w_in]):
            return jsonify({"success": False, "error": "Missing data"}), 400

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Update the specific row
        sql = """
            UPDATE stamps
            SET fingerprint_phash = %s,
                fingerprint_dhash = %s,
                fingerprint_whash = %s
            WHERE id = %s
        """
        cursor.execute(sql, (p_in, d_in, w_in, row_id))
        conn.commit()

        return jsonify({"success": True, "message": f"Row {row_id} updated"})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/get-broken-stamps', methods=['GET'])
def get_broken_stamps():
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        # Use your exact query
        query = "SELECT id, fileName, drive_url FROM stamps WHERE (fingerprint_phash IS NULL OR fingerprint_phash IN ('', 'NO_PHASH')) OR (fingerprint_dhash IS NULL OR fingerprint_dhash IN ('', 'NO_DHASH', 'NO_PHASH')) OR (fingerprint_whash IS NULL OR fingerprint_whash IN ('', 'NO_WHASH', 'NO_PHASH'));"
        cursor.execute(query)
        rows = cursor.fetchall()
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            conn.close()


@app.route('/get-pending-audits', methods=['GET'])
def get_pending_audits():
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    try:
        # We fetch records that are UNRESOLVED and have a valid target image
        # We sort by checked_at DESC so you see the newest finds first
        query = """
            SELECT
                stamp_id,
                original_id,
                status,
                dist_p,
                dist_d,
                dist_w,
                target_fileName,
                target_drive_url,
                original_fileName,
                original_drive_url,
                checked_at
            FROM duplicate_audit
            WHERE user_resolution = 'UNRESOLVED'
              AND status IN ('DUPLICATE', 'UNCERTAIN')
            ORDER BY checked_at DESC
        """

        cursor.execute(query)
        results = cursor.fetchall()

        # Convert timestamps to strings so JSON can handle them
        for row in results:
            if row['checked_at']:
                row['checked_at'] = row['checked_at'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify(results)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/resolve-duplicate', methods=['POST'])
def resolve_duplicate():
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    stamp_id = data.get('stamp_id')
    action = data.get('action')

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    try:
        conn.start_transaction()

        # Fetch Drive URL for storage cleanup
        cursor.execute("SELECT drive_url FROM stamps WHERE id = %s", (stamp_id,))
        record_info = cursor.fetchone()

        if action == 'DELETE_DATA':

            # Cascading delete to maintain DB integrity
            cursor.execute("DELETE FROM stamp_tags WHERE stamp_id = %s", (stamp_id,))
            cursor.execute("DELETE FROM stamp_images WHERE stamp_id = %s", (stamp_id,))
            cursor.execute("DELETE FROM stamps WHERE id = %s", (stamp_id,))

            # Update audit status
            cursor.execute("UPDATE duplicate_audit SET user_resolution = 'DATA_DELETED' WHERE stamp_id = %s",
                           (stamp_id,))

        elif action == 'DELETE_CROP':
            # Path 2: Delete ONLY the image from Google Drive storage
            if record_info and record_info['drive_url']:
                delete_file_from_drive(record_info['drive_url'])

            cursor.execute("UPDATE duplicate_audit SET user_resolution = 'CROP_FILE_DELETED' WHERE stamp_id = %s",
                           (stamp_id,))

        elif action == 'DELETE_DUPLICATE_ENTRY':
            # Path 3: Only remove the audit record from the review list
            cursor.execute("DELETE FROM duplicate_audit WHERE stamp_id = %s", (stamp_id,))

        conn.commit()
        return jsonify({"success": True, "message": f"Action {action} completed successfully."})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/audit-next-stamp', methods=['POST'])
def audit_next_stamp():
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Pick the oldest un-audited stamp
        cursor.execute("""
            SELECT id, fileName, drive_url, fingerprint_phash, fingerprint_dhash, fingerprint_whash
            FROM stamps
            WHERE id NOT IN (SELECT stamp_id FROM duplicate_audit)
            ORDER BY id ASC LIMIT 1
        """)
        target = cursor.fetchone()

        if not target:
            return jsonify({"status": "complete", "message": "All stamps audited."})

        # 2. Duplicate Search Logic - CHANGED TO LIMIT 5
        search_sql = """
            SELECT id, fileName, drive_url,
                   BIT_COUNT(UNHEX(fingerprint_phash) ^ UNHEX(%s)) AS dist_p,
                   BIT_COUNT(UNHEX(fingerprint_dhash) ^ UNHEX(%s)) AS dist_d,
                   BIT_COUNT(UNHEX(fingerprint_whash) ^ UNHEX(%s)) AS dist_w
            FROM stamps
            WHERE id != %s
              AND (fingerprint_phash IS NOT NULL AND fingerprint_phash != '')
            HAVING dist_p <= 25 OR dist_d <= 25 OR dist_w <= 25
            ORDER BY dist_p ASC LIMIT 10
        """
        cursor.execute(search_sql, (target['fingerprint_phash'], target['fingerprint_dhash'],
                                    target['fingerprint_whash'], target['id']))

        # CHANGED: Get all matches up to 5
        matches = cursor.fetchall()

        # 3. Apply Logic to each match
        if matches:
            found_valid_match = False

            for match in matches:
                votes = 0
                if match['dist_p'] <= 12: votes += 1
                if match['dist_d'] <= 12: votes += 1
                if match['dist_w'] <= 12: votes += 1

                confidence = 0
                if target['fileName'] == match['fileName']:
                    confidence = 99
                elif votes == 3:
                    confidence = 99
                elif votes == 2:
                    confidence = 66
                elif votes == 1:
                    confidence = 33

                if confidence > 33:
                    found_valid_match = True
                    status = "DUPLICATE" if confidence >= 66 else "UNCERTAIN"

                    # ADDED IGNORE: Prevents Primary Key crash if same target is matched multiple times
                    insert_sql = """
                        INSERT IGNORE INTO duplicate_audit (
                            stamp_id, original_id, status, dist_p, dist_d, dist_w,
                            target_fileName, target_drive_url,
                            original_fileName, original_drive_url, user_resolution
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'UNRESOLVED')
                    """
                    cursor.execute(insert_sql, (
                        target['id'], match['id'], status,
                        match['dist_p'], match['dist_d'], match['dist_w'],
                        target['fileName'], target['drive_url'],
                        match['fileName'], match['drive_url']
                    ))

            # If none of the 5 matches passed the confidence check, mark target as CLEAN
            if not found_valid_match:
                # ADDED IGNORE here too
                cursor.execute("""
                    INSERT IGNORE INTO duplicate_audit (stamp_id, status, user_resolution)
                    VALUES (%s, 'CLEAN', 'RESOLVED')
                """, (target['id'],))
        else:
            # No fuzzy matches found at all - ADDED IGNORE
            cursor.execute("""
                INSERT IGNORE INTO duplicate_audit (stamp_id, status, user_resolution)
                VALUES (%s, 'CLEAN', 'RESOLVED')
            """, (target['id'],))

        conn.commit()
        return jsonify({
            "status": "success",
            "audited_id": target['id'],
            "matches_found": len(matches) if matches else 0
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/check-cloud-duplicate', methods=['POST'])
def check_cloud_duplicate():
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    sql = ""
    params = []
    executable_sql = ""  # String for console debugging

    try:
        data = request.json
        p_in = data.get('phash')
        d_in = data.get('dhash')
        w_in = data.get('whash')
        incoming_filename = data.get('fileName')

        if not p_in or not d_in or not w_in:
            return jsonify({"success": False, "error": "Missing hashes"}), 400

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # ---------------------------------------------------------
        # 1. EXACT FILENAME CHECK (Priority 1)
        # ---------------------------------------------------------
        if incoming_filename:
            sql_filename = "SELECT * FROM stamps WHERE fileName = %s"
            params_filename = (incoming_filename,)
            cursor.execute(sql_filename, params_filename)
            match_filename = cursor.fetchone()

            if match_filename:
                # Sanitize data (Date/Decimals) before returning
                for key, value in match_filename.items():
                    if isinstance(value, (datetime.date, datetime.datetime)):
                        match_filename[key] = value.isoformat()
                    elif isinstance(value, decimal.Decimal):
                        match_filename[key] = float(value)

                return jsonify({
                    "success": True,
                    "candidates": [],  # Empty as requested
                    "match_id": match_filename['fileName'],
                    "score": 100,
                    "stamp_info": match_filename,  # Return full info
                    "details": {
                        "p_dist": 0,
                        "d_dist": 0,
                        "w_dist": 0,
                        "votes": 3,
                        "method": "Exact Filename Match"
                    }
                })

        # ---------------------------------------------------------
        # 2. STANDARD HASH SEARCH (Priority 2)
        # ---------------------------------------------------------
        sql = """
            SELECT *,
            BIT_COUNT(UNHEX(LPAD(fingerprint_phash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) AS dist_p,
            IFNULL(BIT_COUNT(UNHEX(LPAD(fingerprint_dhash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))), 99) AS dist_d,
            IFNULL(BIT_COUNT(UNHEX(LPAD(fingerprint_whash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))), 99) AS dist_w
            FROM stamps
            WHERE
                (fingerprint_phash IS NOT NULL AND BIT_COUNT(UNHEX(LPAD(fingerprint_phash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) <= 25)
                OR
                (fingerprint_dhash IS NOT NULL AND BIT_COUNT(UNHEX(LPAD(fingerprint_dhash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) <= 25)
                OR
                (fingerprint_whash IS NOT NULL AND BIT_COUNT(UNHEX(LPAD(fingerprint_whash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) <= 25)
            ORDER BY (dist_p + dist_d + dist_w) ASC
            LIMIT 10
        """

        params = (p_in, d_in, w_in, p_in, d_in, w_in)
        executable_sql = sql.replace('%s', "'{}'").format(*params)  # Debug string

        cursor.execute(sql, params)
        matches = cursor.fetchall()

        candidate_results = []
        current_score = 0
        votes = 0

        if matches:
            for row in matches:
                current_score = 0
                votes = 0
                # --- FIX: SANITIZE ROW DATA ---
                # Convert Date/Decimal objects to strings so jsonify doesn't crash
                for key, value in row.items():
                    if isinstance(value, (datetime.date, datetime.datetime)):
                        row[key] = value.isoformat()
                    elif isinstance(value, decimal.Decimal):
                        row[key] = float(value)  # or str(value)

                if row['dist_d'] <= 20:
                    current_score += 33
                    votes += 1

                if row['dist_p'] <= 20:
                    current_score += 33
                    votes += 1

                if row['dist_w'] <= 20:
                    current_score += 33
                    votes += 1

                # --- NEW CONDITION: Only add if score is significantly high ---
                if current_score >= 33:
                    candidate_results.append({
                        "fileName": row['fileName'],
                        "image_vector": row['image_vector'],
                        "score": current_score,
                        "votes": votes,
                        "stamp_info": row,
                        "is_duplicate": current_score >= 33,
                        "debug": {
                            "dist_p": row['dist_p'],
                            "dist_d": row['dist_d'],
                            "dist_w": row['dist_w']
                        }
                    })

            # Sort Highest -> Lowest
            candidate_results.sort(key=lambda x: x['score'], reverse=True)

            return jsonify({
                "success": True,
                "candidates": candidate_results,
                "best_match": candidate_results[0] if candidate_results else None,
                "esql": executable_sql
            })

        return jsonify({"success": False, "candidates": [], "message": "No match found", "esql": executable_sql})

    except Exception as e:
        # Return the error AND the query you can run in your console
        return jsonify({
            "success": False,
            "error_msg": str(e),
            "paste_into_console": executable_sql,
            "debug_info": {
                "raw_sql": sql,
                "applied_params": params,
                "received_lengths": {
                    "p_len": len(str(p_in)) if p_in else 0,
                    "d_len": len(str(d_in)) if d_in else 0,
                    "w_len": len(str(w_in)) if w_in else 0
                }
            }
        }), 200
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


import decimal
import datetime


@app.route('/check-cloud-duplicate-app', methods=['POST'])
def check_cloud_duplicate_app():
    # 1. Auth
    api_key = request.headers.get('X-API-KEY')
    if api_key != SECRET_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    conn = None
    cursor = None
    executable_sql = ""

    try:
        data = request.json
        p_in = data.get('phash', '')
        d_in = data.get('dhash', '')
        w_in = data.get('whash', '')

        # Validation
        if not p_in:
            return jsonify({"is_duplicate": False, "score": 0, "message": "Invalid input"})

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # 2. OPTIMIZED SQL (Single Query)
        sql = """
            SELECT *,
            BIT_COUNT(UNHEX(LPAD(fingerprint_phash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) AS dist_p,
            IFNULL(BIT_COUNT(UNHEX(LPAD(fingerprint_dhash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))), 99) AS dist_d,
            IFNULL(BIT_COUNT(UNHEX(LPAD(fingerprint_whash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))), 99) AS dist_w
            FROM stamps
            WHERE
                (fingerprint_phash IS NOT NULL AND BIT_COUNT(UNHEX(LPAD(fingerprint_phash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) <= 25)
                OR
                (fingerprint_dhash IS NOT NULL AND BIT_COUNT(UNHEX(LPAD(fingerprint_dhash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) <= 25)
                OR
                (fingerprint_whash IS NOT NULL AND BIT_COUNT(UNHEX(LPAD(fingerprint_whash, 16, '0')) ^ UNHEX(LPAD(%s, 16, '0'))) <= 25)
            ORDER BY (dist_p + dist_d + dist_w) ASC
            LIMIT 10
        """

        params = (p_in, d_in, w_in, p_in, d_in, w_in)
        executable_sql = sql.replace('%s', "'{}'").format(*params)  # Debug string

        cursor.execute(sql, params)
        matches = cursor.fetchall()

        candidate_results = []
        current_score = 0
        votes = 0

        if matches:
            for row in matches:
                # --- FIX: SANITIZE ROW DATA ---
                # Convert Date/Decimal objects to strings so jsonify doesn't crash
                for key, value in row.items():
                    if isinstance(value, (datetime.date, datetime.datetime)):
                        row[key] = value.isoformat()
                    elif isinstance(value, decimal.Decimal):
                        row[key] = float(value)  # or str(value)

                # --- SCORING LOGIC ---
                current_score = 0
                votes = 0  # <--- Initialize votes

                if row['dist_d'] <= 20:
                    current_score += 33
                    votes += 1

                if row['dist_p'] <= 20:
                    current_score += 33
                    votes += 1

                if row['dist_w'] <= 20:
                    current_score += 33
                    votes += 1

                # Clean up redundancy:
                # 'row' has everything. We don't need to send image_vector twice.
                # We can construct the clean object:
                # --- NEW CONDITION: Only add if score is significantly high ---
                if current_score >= 33:
                    candidate_results.append({
                        "fileName": row['fileName'],
                        "image_vector": row['image_vector'],
                        "score": current_score,
                        "votes": votes,
                        "stamp_info": row,
                        "is_duplicate": current_score >= 33,
                        "debug": {
                            "dist_p": row['dist_p'],
                            "dist_d": row['dist_d'],
                            "dist_w": row['dist_w']
                        }
                    })

            # Sort Highest -> Lowest
            candidate_results.sort(key=lambda x: x['score'], reverse=True)

            return jsonify({
                "success": True,
                "candidates": candidate_results,
                "best_match": candidate_results[0] if candidate_results else None,
                "esql": executable_sql
            })

        return jsonify({"success": False, "candidates": [], "message": "No match found", "esql": executable_sql})

    except Exception as e:
        print(f"Cloud Check Error: {e}")
        return jsonify({"success": False, "error": str(e), "esql": executable_sql}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
# ONLINE APIS CODE ARE ENDED HERE
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################


# ==========================================
# SERVER RUNNER
# ==========================================
import socket


def get_local_ip():
    try:
        # This creates a quick dummy connection to find your Mac's true Wi-Fi IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


if __name__ == '__main__':
    system_ip = get_local_ip()
    print("\n🚀 Starting Local StampSeva API Server...")
    print(f"🏠 Local Mac URL: http://127.0.0.1:5001")
    print(f"🌍 Network URL:   http://{system_ip}:5001  <-- Use this to connect from other devices!")
    print("--------------------------------------------------\n")

    # host='0.0.0.0' is the magic flag that allows external connections
    app.run(host='0.0.0.0', port=5001, debug=True, use_reloader=False)
    # app.run(port=5001)
